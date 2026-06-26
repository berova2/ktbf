"""
2025_sporcu_listesi.xlsx → lisans.db aktarım scripti
====================================================
- Excel'deki yeni kulüpleri kulupler tablosuna ekler
- Daha önce kayıtlı olmayan sporcuları sporcular tablosuna ekler
- 2025 sezonu için lisans kaydı oluşturur (yeni veya ek lisans)
- Mevcut sporculara 2025 lisansı ekler (kulüp değişikliği varsa)
"""

import openpyxl
import re
import sys
import os
import datetime

sys.path.insert(0, os.path.dirname(__file__))
import lisans_db as db

EXCEL_PATH = r"c:\ktbf\sporcu2025.xlsx"


def temizle(s):
    if s is None:
        return ""
    return str(s).strip()


def tel_temizle(tel):
    if tel is None:
        return None
    s = str(tel).strip()
    try:
        if 'e' in s.lower():
            s = str(int(float(s)))
    except (ValueError, OverflowError):
        pass
    # Biçimlendirme karakterlerini temizle
    s = s.replace("'", "").replace(" ", "").replace("(", "").replace(")", "").replace("-", "").replace("+", "")
    s = s.replace("\u202a", "").replace("\u202c", "").replace("\xa0", "")
    if s.startswith("90") and len(s) > 10:
        s = "0" + s[2:]
    if s and s.isdigit() and len(s) == 12 and s.startswith("90"):
        s = "0" + s[2:]
    if s and s.isdigit() and len(s) == 10:
        s = "0" + s
    return s if s and s.isdigit() else None


def normalize_kulup(ad):
    """2025 Excel'deki kulüp adını normalize eder, varsa eşleşen DB kulübünü döner."""
    ad = temizle(ad).upper().strip()
    
    # Bilinen eşleşmeler
    eslesme = {
        "GREEN PEDAL CYCLING SPOR KULÜBÜ DERNEĞİ": "GREEN PEDAL",
        "GREEN PEDAL CYCLING SPOR KULUBU DERNEĞİ": "GREEN PEDAL",
        "TÜFEKÇI SPOR KULÜBÜ": "TÜFEKÇI",
        "TÜFEKÇİ SPOR KULÜBÜ": "TÜFEKÇI",
        "TÜFEKÇI SPOR KULUBU": "TÜFEKÇI",
        "VELO CYPRI KULÜBÜ DERNEĞİ": "VELO CYPRI",
        "VELO CYPRI KULUBU DERNEĞİ": "VELO CYPRI",
        "NECATI HASAN GÜNEY CYCLING TEAM SPOR KULÜBÜ DERNEĞİ": "NGH",
        "NECATI HASAN GÜNEY CYCLING TEAM SPOR KULUBU DERNEĞİ": "NGH",
        "NECATI HASAN GUNEY CYCLING TEAM SPOR KULUBU DERNEĞİ": "NGH",
        "KKTC BISIKLET SEVENLER DERNEĞİ": "KKTC BİSİKLET SEVENLER DERNEĞİ",
        "ISLANDERS SPORTS CLUB ADALILAR SPOR DERNEĞİ": "ISLANDERS SPORTS CLUB ADALILAR SPOR KULÜBÜ DERNEĞİ",
        "VELOSPEED SPOR DERNEĞİ": "VELOSPEED SPOR DERNEĞİ",
        "DAĞ TENİS KULÜBÜ": "DAĞ TENİS KULÜBÜ",
        "DAĞ TENIS KULUBU": "DAĞ TENİS KULÜBÜ",
    }
    
    return eslesme.get(ad, ad)


def main():
    print("=" * 70)
    print("2025 Sporcu Listesi → Veritabanı aktarımı")
    print("=" * 70)

    db.init_db()

    # Mevcut veritabanı verilerini yükle
    with db.get_conn() as conn:
        mevcut_sporcular = conn.execute("""
            SELECT s.id, s.ad, s.soyad, s.kimlik_no, s.dogum_tarihi,
                   l.lisans_no, l.sezon, k.ad as kulup_adi, l.id as lisans_id
            FROM sporcular s
            LEFT JOIN lisanslar l ON l.sporcu_id = s.id
            LEFT JOIN kulupler k ON k.id = l.kulup_id
            ORDER BY s.id
        """).fetchall()
        
        mevcut_kulupler = conn.execute(
            "SELECT id, ad FROM kulupler"
        ).fetchall()

    # Veritabanı indeksleri
    lisans_no_db = {}
    ad_soyad_db = {}
    for r in mevcut_sporcular:
        if r['lisans_no']:
            lisans_no_db[str(r['lisans_no']).strip()] = r
        key = (r['ad'].strip().upper() if r['ad'] else '',
               r['soyad'].strip().upper() if r['soyad'] else '',
               str(r['dogum_tarihi'] or '').strip())
        ad_soyad_db[key] = r
    
    kulup_adi_db = {r['ad'].upper().strip(): r for r in mevcut_kulupler}

    # 2025 Excel'i oku
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Worksheet"]
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    print(f"\n2025 Excel'de {len(rows)} satır veri bulundu.")

    # ---- 1. Yeni kulüpleri bul ve ekle ----
    yeni_kulup_set = set()
    for row in rows:
        kulup_ham = row[17]  # R sütunu (0-indexed: 17) = Kulübü
        if kulup_ham:
            normalized = normalize_kulup(str(kulup_ham))
            if normalized not in kulup_adi_db:
                yeni_kulup_set.add(normalized)

    print(f"\nYeni kulüp adayları: {len(yeni_kulup_set)}")
    for yk in sorted(yeni_kulup_set):
        kid = db.kulup_ekle(ad=yk, sezon="2025", durum="Aktif")
        kulup_adi_db[yk.upper()] = {"id": kid, "ad": yk}
        print(f"  ✅ Yeni kulüp eklendi: {yk} (id={kid})")

    # ---- 2. Sporcuları işle ----
    eklenen_sporcu = 0
    eklenen_lisans = 0
    atlanan = 0
    hata = 0

    print(f"\nSporcular işleniyor...\n")

    for idx, row in enumerate(rows, start=2):
        isim = temizle(row[2])      # C: İsim
        soyisim = temizle(row[3])   # D: Soyisim
        telefon = row[4]            # E: Telefon
        email = temizle(row[5])     # F: Email
        lisans_no_str = temizle(str(row[6])) if row[6] is not None else ""  # G: Lisans No
        uyruk = temizle(row[7])     # H: Uyruk
        dogum_tarihi = temizle(row[9])  # J: Doğum Tarihi (YYYY-MM-DD)
        cinsiyet = temizle(row[14]) # O: Cinsiyet
        kulup_ham = temizle(row[17]) # R: Kulübü
        baba_adi = temizle(row[13]) if row[13] else None  # N: Baba Adı

        if not isim or not soyisim:
            atlanan += 1
            continue

        # Cinsiyet
        if cinsiyet.upper() == "KADIN":
            cins = "Kadın"
        elif cinsiyet.upper() in ("ERKEK", "ERKEK"):
            cins = "Erkek"
        else:
            cins = "Belirtilmedi"

        # Uyruk
        uyruk_std = uyruk.upper()
        if uyruk_std in ("KKTC", "K.K.T.C.", "K.K.T.C", "KIBRIS", "KIBRISLI", "KIBRISLİ", "KTCC"):
            uyruk_std = "KKTC"
        elif uyruk_std in ("TC", "T.C.", "T.C", "TÜRK", "TURK"):
            uyruk_std = "T.C."
        elif uyruk_std in ("INGILIZ", "İNGİLİZ", "INGILTERE", "İNGİLTERE"):
            uyruk_std = "İngiltere"
        else:
            uyruk_std = uyruk_std or "KKTC"

        # Telefon
        telefon_temiz = tel_temizle(telefon)

        # Kulüp
        kulup_adi_normalized = normalize_kulup(kulup_ham) if kulup_ham else None
        kulup_id = None
        if kulup_adi_normalized:
            k_row = kulup_adi_db.get(kulup_adi_normalized.upper())
            if k_row:
                kulup_id = k_row["id"]

        # Lisans türü
        lisans_turu = "Ferdi" if kulup_id is None else "Ulusal"

        # Mevcut kaydı kontrol et: önce lisans_no ile
        mevcut = None
        if lisans_no_str and lisans_no_str.replace(" ", "").isdigit():
            temiz_lisans_no = str(int(float(lisans_no_str)))
        else:
            temiz_lisans_no = lisans_no_str
        
        mevcut = lisans_no_db.get(temiz_lisans_no)
        
        if not mevcut:
            # Alternatif: lisans_no'nun son kısmıyla dene (bazıları "2025-059" formatında)
            if temiz_lisans_no:
                # Tam eşleşme yoksa, 2020'deki lisans numaralarıyla dene
                for ln_key, sp in lisans_no_db.items():
                    # 2020 lisans numaraları 7 haneli (3081076 vb)
                    # 2025 lisans numaraları "2025-059" veya kısa format
                    pass
            
            # Ad+Soyad+Doğum ile dene
            dogum_key = str(dogum_tarihi or "").strip()
            key = (isim.upper(), soyisim.upper(), dogum_key)
            if key in ad_soyad_db:
                mevcut = ad_soyad_db[key]
            else:
                # Sadece ad+soyad ile dene (doğum tarihi olmayanlar için)
                for k, v in ad_soyad_db.items():
                    if k[0] == isim.upper() and k[1] == soyisim.upper():
                        mevcut = v
                        break

        if mevcut:
            # Mevcut sporcu - 2025 lisansı var mı kontrol et
            sporcu_id = mevcut["id"]
            with db.get_conn() as conn:
                var_2025 = conn.execute(
                    "SELECT id FROM lisanslar WHERE sporcu_id=? AND sezon='2025'",
                    (sporcu_id,)
                ).fetchone()
            
            if var_2025:
                print(f"  ⏩ {isim} {soyisim} → 2025 lisansı zaten var")
                atlanan += 1
            else:
                # 2025 lisansı ekle
                try:
                    db.sporcu_guncelle(sporcu_id,
                        telefon=telefon_temiz,
                        email=email or None,
                    )
                    db.lisans_ekle(
                        sporcu_id=sporcu_id,
                        lisans_no=f"2025-{sporcu_id:03d}",
                        lisans_turu=lisans_turu,
                        sezon="2025",
                        kulup_id=kulup_id,
                    )
                    eklenen_lisans += 1
                    print(f"  ➕ {isim} {soyisim} → 2025 lisansı eklendi (kulüp: {kulup_adi_normalized or 'Ferdi'})")
                except Exception as e:
                    hata += 1
                    print(f"  ❌ HATA (lisans): {isim} {soyisim} → {e}")
        else:
            # Yeni sporcu
            try:
                kimlik_no = temiz_lisans_no if temiz_lisans_no else f"2025-{idx:03d}"
                
                sporcu_id = db.sporcu_ekle(
                    ad=isim,
                    soyad=soyisim,
                    kimlik_no=kimlik_no,
                    dogum_tarihi=dogum_tarihi or None,
                    cinsiyet=cins,
                    uyruk=uyruk_std,
                    telefon=telefon_temiz,
                    email=email or None,
                    spor_dairesi_kayitli=0,
                )
                
                db.lisans_ekle(
                    sporcu_id=sporcu_id,
                    lisans_no=temiz_lisans_no or f"2025-{sporcu_id:03d}",
                    lisans_turu=lisans_turu,
                    sezon="2025",
                    kulup_id=kulup_id,
                )
                eklenen_sporcu += 1
                print(f"  ✅ YENİ: {isim} {soyisim} → sporcu(id={sporcu_id}), kulüp: {kulup_adi_normalized or 'Ferdi'}")
            except Exception as e:
                hata += 1
                print(f"  ❌ HATA (yeni sporcu): {isim} {soyisim} → {e}")

    print("\n" + "=" * 70)
    print(f"Özet:")
    print(f"  Yeni sporcu     : {eklenen_sporcu}")
    print(f"  Yeni 2025 lisans: {eklenen_lisans}")
    print(f"  Atlanan         : {atlanan}")
    print(f"  Hata            : {hata}")
    print("=" * 70)


if __name__ == "__main__":
    main()
