"""
2020_sporcu_kütüğü.xlsx → lisans.db aktarım scripti
======================================================
- Excel'deki kulüp isimlerini kulupler tablosuna ekler
- Sporcu bilgilerini sporcular tablosuna ekler
- Her sporcu için ilgili kulübe bağlı lisans kaydı oluşturur
"""

import openpyxl
import re
import sys
import os

# Proje ana dizinini ekle
sys.path.insert(0, os.path.dirname(__file__))
import lisans_db as db

EXCEL_PATH = r"c:\ktbf\sporcu.xlsx"

# "FERDİ" bireysel lisans anlamına gelir, kulüp sayılmaz
BIREYSEL = {"FERDİ", "FERDİ", "FERDI", "FERDİ"}

def temizle_ve_normalize(s):
    """Metni temizler ve None yerine boş string döner."""
    if s is None:
        return ""
    s = str(s).strip()
    return s


def ad_soyad_ayir(ad_soyad: str):
    """
    'Adı ve Soyadı' alanını ad ve soyad olarak ayırır.
    Son kelime soyaddır, kalanı addır.
    """
    ad_soyad = temizle_ve_normalize(ad_soyad)
    if not ad_soyad:
        return "", ""
    parts = ad_soyad.split()
    if len(parts) == 1:
        return parts[0], ""
    soyad = parts[-1]
    ad = " ".join(parts[:-1])
    return ad, soyad


def dogum_tarihini_coz(deger):
    """
    Doğum tarihi sütunundaki değeri çözer.
    Çeşitli formatları dener:
    - datetime objesi
    - '07.11.1982 / Lefkoşa' gibi
    - '29Nisan1980 Ortaköy' gibi
    - '01/05/1981 G.Mağusa' gibi
    - 'Yenisehir/lefkosa-14.06.1996' gibi
    """
    if deger is None:
        return None, None

    import datetime

    if isinstance(deger, datetime.datetime):
        return deger.strftime("%Y-%m-%d"), None

    s = str(deger).strip()

    # Tarih ve yer bilgisini ayır
    # Önce '-' ile ayrılmış olabilir: "Yenisehir/lefkosa-14.06.1996"
    # Veya '/' ile ayrılmış: "07.11.1982 / Lefkoşa"
    # Veya boşlukla ayrılmış: "29Nisan1980 Ortaköy"

    # Yaygın format: "dd.mm.yyyy / yer" veya "dd.mm.yyyy   yer"
    # Tarih kısmını ayıklamaya çalış
    tarih_str = s
    yer = None

    # Pattern 1: "dd.mm.yyyy / yer" veya "dd.mm.yyyy yer"
    m = re.match(r'(\d{1,2})[./](\d{1,2})[./](\d{4})\s*[/\-\s]*\s*(.*)', s)
    if m:
        gun, ay, yil, yer = m.groups()
        try:
            dt = datetime.date(int(yil), int(ay), int(gun))
            return dt.strftime("%Y-%m-%d"), yer.strip() if yer.strip() else None
        except ValueError:
            pass

    # Pattern 2: "gg.aa.yyyy-yer" (tire ile ayrılmış, yer önce)
    m = re.match(r'(.+?)[-/](\d{1,2})[./](\d{1,2})[./](\d{4})', s)
    if m:
        yer, gun, ay, yil = m.groups()
        try:
            dt = datetime.date(int(yil), int(ay), int(gun))
            return dt.strftime("%Y-%m-%d"), yer.strip() if yer.strip() else None
        except ValueError:
            pass

    # Pattern 3: "GunAyAdıYıl yer" (Türkçe ay adları ile)
    aylar = {
        'ocak': 1, 'şubat': 2, 'mart': 3, 'nisan': 4, 'mayıs': 5,
        'mayis': 5, 'haziran': 6, 'temmuz': 7, 'ağustos': 8,
        'agustos': 8, 'eylül': 9, 'eylul': 9, 'ekim': 10,
        'kasım': 11, 'kasim': 11, 'aralık': 12, 'aralik': 12
    }
    for ay_adi, ay_no in aylar.items():
        m = re.search(r'(\d{1,2})\s*' + ay_adi + r'\s*(\d{4})', s, re.IGNORECASE)
        if m:
            gun, yil = int(m.group(1)), int(m.group(2))
            try:
                dt = datetime.date(yil, ay_no, gun)
                # Yer bilgisini ayıkla
                yer_kismi = s[:m.start()] + s[m.end():]
                yer = yer_kismi.strip().strip('/-').strip() or None
                return dt.strftime("%Y-%m-%d"), yer
            except ValueError:
                pass

    # Pattern 4: Sadece tarih "dd.mm.yyyy" veya "dd/mm/yyyy"
    m = re.match(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', s)
    if m:
        gun, ay, yil = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            dt = datetime.date(yil, ay, gun)
            return dt.strftime("%Y-%m-%d"), None
        except ValueError:
            pass

    # Pattern 5: "Ay.yıl" gibi eksik format ("0.06.2005")
    m = re.match(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', s)
    if m:
        gun, ay, yil = m.groups()
        try:
            dt = datetime.date(int(yil), int(ay), int(gun or 1))
            return dt.strftime("%Y-%m-%d"), None
        except ValueError:
            pass

    return None, None


def telefon_temizle(tel):
    """Telefon numarasını temizler, başındaki +90 veya 0'ı düzenler."""
    if tel is None:
        return None
    s = str(tel).strip()
    # Bilimsel notasyondan kurtul (örn: 5.48870542856534e+20)
    try:
        if 'e' in s.lower():
            s = str(int(float(s)))
    except (ValueError, OverflowError):
        pass
    s = s.replace("'", "").replace(" ", "")
    s = s.replace("+90", "0")
    if s and s.isdigit() and len(s) > 11:
        s = s[-11:]  # Son 11 haneyi al
    if s and s.isdigit() and len(s) == 12 and s.startswith("90"):
        s = "0" + s[2:]
    if s and not s.startswith("0") and s.isdigit() and len(s) == 10:
        s = "0" + s
    return s if s else None


def normalize_kulup_adi(ad):
    """Kulüp adını normalize eder."""
    ad = temizle_ve_normalize(ad).upper()
    return ad


def main():
    print("=" * 60)
    print("Excel → Veritabanı aktarımı başlıyor")
    print("=" * 60)

    # Veritabanını başlat
    db.init_db()

    # Excel'i oku
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sayfa1"]

    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    print(f"Excel'de {len(rows)} satır veri bulundu.")

    # ---- 1. Benzersiz kulüpleri bul ----
    kulup_set = set()
    for row in rows:
        kulup_adi = row[14]  # Column 15 (0-indexed: 14) = Kulübü
        kulup_adi = temizle_ve_normalize(kulup_adi).upper()
        if kulup_adi and kulup_adi not in BIREYSEL:
            kulup_set.add(kulup_adi)

    print(f"\nBenzersiz kulüp sayısı: {len(kulup_set)}")
    print("Kulüpler:", sorted(kulup_set))

    # ---- 2. Kulüpleri veritabanına ekle ----
    kulup_map = {}  # {kulup_adi: kulup_id}
    for k_adi in sorted(kulup_set):
        # Daha önce eklenmiş mi kontrol et
        var_kulup = None
        with db.get_conn() as conn:
            var_kulup = conn.execute(
                "SELECT id FROM kulupler WHERE ad=?", (k_adi,)
            ).fetchone()
        if var_kulup:
            kulup_map[k_adi] = var_kulup["id"]
            print(f"  ⏩ Kulüp zaten var: {k_adi} (id={var_kulup['id']})")
        else:
            kid = db.kulup_ekle(ad=k_adi, sezon="2020", durum="Aktif")
            kulup_map[k_adi] = kid
            print(f"  ✅ Kulüp eklendi: {k_adi} (id={kid})")

    # ---- 3. Sporcuları ekle ----
    eklenen = 0
    atlanan = 0
    hata = 0

    print(f"\nSporcular ekleniyor...")

    for idx, row in enumerate(rows, start=2):
        ad_soyad = row[0]   # Adı ve Soyadı
        lisans_no = row[1]  # Lisans No
        kimlik_no = row[2]  # Kimlik No
        email = row[4]      # Email
        adres = row[5]      # Adres
        cinsiyet = row[6]   # Cinsiyet
        dogum_str = row[7]  # Doğum Tarihi ve Yeri
        uyruk = row[11]     # Tabiyeti (uyruk)
        ev_tel = row[12]    # Ev Telefonu
        cep_tel = row[13]   # Cep Telefonu
        kulup_adi = row[14] # Kulübü

        ad_soyad = temizle_ve_normalize(ad_soyad)
        if not ad_soyad:
            atlanan += 1
            continue

        # Ad/Soyad ayır
        ad, soyad = ad_soyad_ayir(ad_soyad)

        # Cinsiyet
        cins = temizle_ve_normalize(cinsiyet).upper()
        if cins == "KADIN":
            cins = "Kadın"
        elif cins == "ERKEK":
            cins = "Erkek"
        else:
            cins = "Belirtilmedi"

        # Uyruk
        uyruk_str = temizle_ve_normalize(uyruk).upper()
        if uyruk_str in ("KKTC", "K.K.T.C.", "K.K.T.C", "KIBRISLI", "KIBRISLİ", "KİBRİSLİ", "KİBRİSLİ"):
            uyruk_str = "KKTC"
        elif uyruk_str in ("T.C", "T.C.", "TÜRK", "TURK"):
            uyruk_str = "T.C."
        else:
            uyruk_str = uyruk_str or "KKTC"

        # Doğum tarihi
        dogum_tarihi, dogum_yeri = dogum_tarihini_coz(dogum_str)

        # Telefon
        telefon = telefon_temizle(cep_tel) or telefon_temizle(ev_tel)

        # Kimlik no
        kimlik = temizle_ve_normalize(str(kimlik_no)) if kimlik_no else ""
        # Sayısal değilse veya çok uzunsa lisans_no kullan
        if kimlik and not kimlik.isdigit():
            kimlik = ""
        if len(kimlik) > 20:
            kimlik = ""

        # Sporcu veritabanında var mı kontrol et (lisans_no üzerinden)
        mevcut = None
        if lisans_no:
            with db.get_conn() as conn:
                # Önce lisanslar tablosundan sporcu_id bul
                lisans_row = conn.execute(
                    """SELECT l.sporcu_id FROM lisanslar l
                       WHERE l.lisans_no=?""", (str(lisans_no),)
                ).fetchone()
                if lisans_row:
                    mevcut = lisans_row["sporcu_id"]

        if mevcut:
            atlanan += 1
            print(f"  ⏩ Atlanıyor (zaten var): {ad_soyad}")
            continue

        try:
            # Sporcu ekle
            sporcu_id = db.sporcu_ekle(
                ad=ad,
                soyad=soyad,
                kimlik_no=kimlik or f"EXCEL-{idx:03d}",
                dogum_tarihi=dogum_tarihi,
                cinsiyet=cins,
                uyruk=uyruk_str,
                telefon=telefon,
                email=temizle_ve_normalize(email) or None,
                adres=temizle_ve_normalize(adres) or None,
                spor_dairesi_kayitli=0
            )

            # Lisans türü ve kulüp ID
            kulup_adi_normalized = temizle_ve_normalize(kulup_adi).upper()
            if kulup_adi_normalized in BIREYSEL or not kulup_adi_normalized:
                lisans_turu = "Ferdi"
                kulup_id = None
            else:
                lisans_turu = "Ulusal"
                kulup_id = kulup_map.get(kulup_adi_normalized)

            # Lisans kaydı oluştur
            db.lisans_ekle(
                sporcu_id=sporcu_id,
                lisans_no=str(lisans_no) if lisans_no else None,
                lisans_turu=lisans_turu,
                sezon="2020",
                kulup_id=kulup_id,
            )

            eklenen += 1
            print(f"  ✅ {ad_soyad} → sporcu(id={sporcu_id}), lisans_turu={lisans_turu}")

        except Exception as e:
            hata += 1
            print(f"  ❌ HATA: {ad_soyad} → {e}")

    print("\n" + "=" * 60)
    print(f"Özet: {eklenen} sporcu eklendi, {atlanan} atlandı, {hata} hata")
    print("=" * 60)


if __name__ == "__main__":
    main()
