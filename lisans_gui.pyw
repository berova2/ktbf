"""
KKTC Bisiklet Federasyonu – Kulüp ve Sporcu Kayıt Arayüzü
==========================================================
Kullanım: python lisans_gui.py
"""

import os
import sys
import ctypes
import subprocess
import time
from datetime import date
import tempfile
import tkinter as tk
from tkinter import ttk, messagebox
import urllib.request
import webbrowser
import lisans_db as db

# ---------------------------------------------------------------------------
# Renk / stil sabitleri
# ---------------------------------------------------------------------------
BG       = "#f4f6f9"
HEADER   = "#1a3c6e"
ACCENT   = "#2563eb"
ROW_ODD  = "#ffffff"
ROW_EVEN = "#eef2f8"
BTN_ADD  = "#16a34a"
BTN_UPD  = "#d97706"
BTN_DEL  = "#dc2626"

BASE_SCREEN_WIDTH = 1920
BASE_SCREEN_HEIGHT = 1080
UI_SCALE = 1.0
FONT     = ("Segoe UI", 10)
FONT_B   = ("Segoe UI", 10, "bold")
FONT_H   = ("Segoe UI", 13, "bold")
FONT_S   = ("Segoe UI", 8)


def _scaled(value: int) -> int:
    return max(1, int(round(value * UI_SCALE)))


def _update_ui_metrics(scale: float) -> None:
    global UI_SCALE, FONT, FONT_B, FONT_H, FONT_S

    UI_SCALE = max(0.9, min(scale, 1.35))
    body_size = _scaled(10)
    header_size = _scaled(13)
    small_size = _scaled(8)
    FONT = ("Segoe UI", body_size)
    FONT_B = ("Segoe UI", body_size, "bold")
    FONT_H = ("Segoe UI", header_size, "bold")
    FONT_S = ("Segoe UI", small_size)


def _apply_window_geometry(window: tk.Misc,
                           width_ratio: float = 0.88,
                           height_ratio: float = 0.85,
                           min_width: int = 1100,
                           min_height: int = 760) -> None:
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    width = min(max(int(screen_width * width_ratio), min_width), screen_width - 40)
    height = min(max(int(screen_height * height_ratio), min_height), screen_height - 60)
    x_pos = max((screen_width - width) // 2, 0)
    y_pos = max((screen_height - height) // 2, 0)
    window.geometry(f"{width}x{height}+{x_pos}+{y_pos}")


def _configure_display(root: tk.Tk) -> None:
    scale = min(
        root.winfo_screenwidth() / BASE_SCREEN_WIDTH,
        root.winfo_screenheight() / BASE_SCREEN_HEIGHT,
    )
    _update_ui_metrics(scale)
    _apply_window_geometry(root)
    if os.name == "nt":
        try:
            root.state("zoomed")
        except tk.TclError:
            pass


def _ensure_tcl_tk_env() -> None:
    """Windows kurulumunda eksikse Tcl/Tk kütüphane yollarını ayarlar."""
    if os.name != "nt":
        return
    if os.environ.get("TCL_LIBRARY") and os.environ.get("TK_LIBRARY"):
        return

    tcl_root = os.path.join(sys.base_prefix, "tcl")
    tcl_dir = os.path.join(tcl_root, "tcl8.6")
    tk_dir = os.path.join(tcl_root, "tk8.6")
    if os.path.isdir(tcl_dir) and os.path.isdir(tk_dir):
        os.environ.setdefault("TCL_LIBRARY", tcl_dir)
        os.environ.setdefault("TK_LIBRARY", tk_dir)


def _enable_windows_dpi_awareness() -> None:
    if os.name != "nt":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


def _style_widget(root: tk.Tk) -> ttk.Style:
    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure("TFrame",       background=BG)
    style.configure("TLabel",       background=BG, font=FONT)
    style.configure("TEntry",       font=FONT, padding=_scaled(3))
    style.configure("TCombobox",    font=FONT)
    style.configure("Header.TLabel", background=HEADER, foreground="white",
                    font=FONT_H, padding=_scaled(5))
    style.configure("Treeview",      font=FONT, rowheight=_scaled(26))
    style.configure("Treeview.Heading", font=FONT_B,
                    background=HEADER, foreground="white")
    style.map("Treeview.Heading",
              background=[("active", ACCENT)])
    style.configure("Add.TButton",   font=FONT_B, background=BTN_ADD,
                    foreground="white", padding=_scaled(5))
    style.configure("Upd.TButton",   font=FONT_B, background=BTN_UPD,
                    foreground="white", padding=_scaled(5))
    style.configure("Del.TButton",   font=FONT_B, background=BTN_DEL,
                    foreground="white", padding=_scaled(5))
    style.configure("Neu.TButton",   font=FONT_B, background=ACCENT,
                    foreground="white", padding=_scaled(5))
    return style


# ---------------------------------------------------------------------------
# Yardımcı widgetlar
# ---------------------------------------------------------------------------

def _lbl_entry(parent, text, row, col=0, width=22, colspan=1):
    ttk.Label(parent, text=text).grid(row=row, column=col,
                        sticky="e", padx=(_scaled(8), _scaled(4)), pady=_scaled(4))
    var = tk.StringVar()
    e = ttk.Entry(parent, textvariable=var, width=width)
    e.grid(row=row, column=col + 1, columnspan=colspan,
        sticky="ew", padx=(0, _scaled(8)), pady=_scaled(4))
    return var


def _lbl_combo(parent, text, values, row, col=0, width=22):
    ttk.Label(parent, text=text).grid(row=row, column=col,
                                       sticky="e", padx=(_scaled(8), _scaled(4)), pady=_scaled(4))
    var = tk.StringVar()
    cb = ttk.Combobox(parent, textvariable=var, values=values,
                      width=width - 2, state="readonly")
    cb.grid(row=row, column=col + 1, sticky="ew", padx=(0, _scaled(8)), pady=_scaled(4))
    return var


def _lbl_check(parent, text, row, col=0):
    var = tk.IntVar()
    chk = ttk.Checkbutton(parent, text=text, variable=var)
    chk.grid(row=row, column=col, columnspan=2,
             sticky="w", padx=(_scaled(60), _scaled(8)), pady=_scaled(2))
    return var


def _make_tree(parent, columns: list[tuple[str, str, int]]):
    """columns: [(id, başlık, genişlik)]"""
    frame = ttk.Frame(parent)
    vsb = ttk.Scrollbar(frame, orient="vertical")
    hsb = ttk.Scrollbar(frame, orient="horizontal")
    tree = ttk.Treeview(frame,
                        columns=[c[0] for c in columns],
                        show="headings",
                        yscrollcommand=vsb.set,
                        xscrollcommand=hsb.set,
                        selectmode="browse")
    vsb.config(command=tree.yview)
    hsb.config(command=tree.xview)
    for cid, ctitle, cw in columns:
        tree.heading(cid, text=ctitle)
        tree.column(cid, width=_scaled(cw), minwidth=_scaled(40))
    vsb.pack(side="right", fill="y")
    hsb.pack(side="bottom", fill="x")
    tree.pack(fill="both", expand=True)
    tree.tag_configure("odd",  background=ROW_ODD)
    tree.tag_configure("even", background=ROW_EVEN)
    tree.tag_configure("group", background=HEADER, foreground="white",
                       font=FONT_B)
    return frame, tree


def _fill_tree(tree, rows):
    tree.delete(*tree.get_children())
    for i, row in enumerate(rows):
        tag = "even" if i % 2 == 0 else "odd"
        tree.insert("", "end", iid=str(row[0]),
                    values=list(row), tags=(tag,))


# ---------------------------------------------------------------------------
# KULÜP sekmesi
# ---------------------------------------------------------------------------

class KulupSekme(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, style="TFrame")
        self._secili_id = None
        self._build()
        self._listele()

    def _build(self):
        ttk.Label(self, text="KULÜP KAYIT VE YÖNETİMİ",
                  style="Header.TLabel").pack(fill="x")

        pane = ttk.PanedWindow(self, orient="horizontal")
        pane.pack(fill="both", expand=True, padx=8, pady=8)

        # Sol: form
        form_frame = ttk.LabelFrame(pane, text="Kulüp Bilgileri", padding=8)
        pane.add(form_frame, weight=1)

        self.v_ad      = _lbl_entry(form_frame, "Kulüp Adı *",    0)
        self.v_yetkili = _lbl_entry(form_frame, "Yetkili Adı",    1)
        self.v_tel     = _lbl_entry(form_frame, "Telefon",        2)
        self.v_email   = _lbl_entry(form_frame, "E-posta",        3)
        self.v_adres   = _lbl_entry(form_frame, "Adres",          4)
        self.v_renk    = _lbl_entry(form_frame, "Forma Rengi",    5)
        self.v_sezon   = _lbl_entry(form_frame, "Sezon",          6, width=10)
        self.v_sezon.set("2026")
        self.v_durum   = _lbl_combo(form_frame, "Durum",
                                    ["Aktif", "Pasif", "Askıda"],  7, width=14)
        self.v_durum.set("Aktif")
        self.v_aidat   = _lbl_check(form_frame, "Aidat Ödendi",   8)

        form_frame.columnconfigure(1, weight=1)

        btn = ttk.Frame(form_frame)
        btn.grid(row=9, column=0, columnspan=2, pady=(12, 4))
        ttk.Button(btn, text="➕ Kaydet",  style="Add.TButton",
                   command=self._kaydet).pack(side="left", padx=4)
        ttk.Button(btn, text="✏️ Güncelle", style="Upd.TButton",
                   command=self._guncelle).pack(side="left", padx=4)
        ttk.Button(btn, text="🗑 Sil", style="Del.TButton",
               command=self._sil).pack(side="left", padx=4)
        ttk.Button(btn, text="🔄 Temizle", style="Neu.TButton",
                   command=self._temizle).pack(side="left", padx=4)
        ttk.Button(btn, text="🧾 Kayıt Formu Üret", style="Neu.TButton",
               command=self._kulup_form_onizleme_ac).pack(side="left", padx=4)

        # Sağ: liste
        list_frame = ttk.LabelFrame(pane, text="Kulüp Listesi", padding=4)
        pane.add(list_frame, weight=2)

        srch_frame = ttk.Frame(list_frame)
        srch_frame.pack(fill="x", pady=(0, 4))
        ttk.Label(srch_frame, text="Durum filtresi:").pack(side="left", padx=4)
        self.v_filtre = tk.StringVar(value="Tümü")
        cb = ttk.Combobox(srch_frame, textvariable=self.v_filtre,
                          values=["Tümü", "Aktif", "Pasif", "Askıda"],
                          width=10, state="readonly")
        cb.pack(side="left")
        cb.bind("<<ComboboxSelected>>", lambda _: self._listele())

        cols = [("id","ID",40), ("ad","Kulüp Adı",180),
                ("yetkili_adi","Yetkili",130), ("telefon","Tel",110),
                ("sezon","Sezon",55), ("durum","Durum",65),
                ("aidat_odendi","Aidat",50)]
        tf, self.tree = _make_tree(list_frame, cols)
        tf.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_sec)

    def _listele(self):
        f = self.v_filtre.get()
        rows = db.kulupler_listele(None if f == "Tümü" else f)
        _fill_tree(self.tree, rows)

    def _on_sec(self, _=None):
        sel = self.tree.selection()
        if not sel:
            return
        self._secili_id = int(sel[0])
        row = db.kulup_getir(self._secili_id)
        if row:
            self.v_ad.set(row["ad"] or "")
            self.v_yetkili.set(row["yetkili_adi"] or "")
            self.v_tel.set(row["telefon"] or "")
            self.v_email.set(row["email"] or "")
            self.v_adres.set(row["adres"] or "")
            self.v_renk.set(row["forma_renk"] or "")
            self.v_sezon.set(row["sezon"] or "")
            self.v_durum.set(row["durum"] or "Aktif")
            self.v_aidat.set(row["aidat_odendi"] or 0)

    def _kaydet(self):
        if not self.v_ad.get().strip():
            messagebox.showwarning("Uyarı", "Kulüp Adı zorunludur.")
            return
        db.kulup_ekle(
            self.v_ad.get().strip(),
            yetkili_adi=self.v_yetkili.get() or None,
            telefon=self.v_tel.get() or None,
            adres=self.v_adres.get() or None,
            email=self.v_email.get() or None,
            forma_renk=self.v_renk.get() or None,
            sezon=self.v_sezon.get() or None,
            durum=self.v_durum.get(),
            aidat_odendi=self.v_aidat.get(),
        )
        self._temizle()
        self._listele()
        messagebox.showinfo("Başarılı", "Kulüp kaydedildi.")

    def _guncelle(self):
        if not self._secili_id:
            messagebox.showwarning("Uyarı", "Listeden bir kulüp seçin.")
            return
        db.kulup_guncelle(
            self._secili_id,
            ad=self.v_ad.get().strip(),
            yetkili_adi=self.v_yetkili.get() or None,
            telefon=self.v_tel.get() or None,
            adres=self.v_adres.get() or None,
            email=self.v_email.get() or None,
            forma_renk=self.v_renk.get() or None,
            sezon=self.v_sezon.get() or None,
            durum=self.v_durum.get(),
            aidat_odendi=self.v_aidat.get(),
        )
        self._listele()
        messagebox.showinfo("Başarılı", "Kulüp güncellendi.")

    def _sil(self):
        if not self._secili_id:
            messagebox.showwarning("Uyarı", "Listeden bir kulüp seçin.")
            return
        if not messagebox.askyesno("Onay", "Seçili kulüp silinsin mi?"):
            return
        try:
            db.kulup_sil(self._secili_id)
            self._temizle()
            self._listele()
            messagebox.showinfo("Başarılı", "Kulüp silindi.")
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    def _temizle(self):
        for v in (self.v_ad, self.v_yetkili, self.v_tel,
                  self.v_email, self.v_adres, self.v_renk):
            v.set("")
        self.v_sezon.set("2026")
        self.v_durum.set("Aktif")
        self.v_aidat.set(0)
        self._secili_id = None
        self.tree.selection_remove(*self.tree.selection())

    def _kulup_form_metni_olustur(self) -> str:
        kulup_adi = self.v_ad.get().strip()
        if not kulup_adi:
            raise ValueError("Form üretmek için önce Kulüp Adı girin.")

        bugun = date.today().strftime("%d.%m.%Y")
        satirlar = [
            "KIBRIS TÜRK BİSİKLET FEDERASYONU",
            "KULÜP YENİDEN KAYIT / AKTİF ÜYELİK BAŞVURU FORMU",
            "=" * 64,
            f"Tarih              : {bugun}",
            f"Kulüp Adı          : {kulup_adi}",
            f"Yetkili Adı        : {self.v_yetkili.get().strip() or '—'}",
            f"Telefon            : {self.v_tel.get().strip() or '—'}",
            f"E-posta            : {self.v_email.get().strip() or '—'}",
            f"Adres              : {self.v_adres.get().strip() or '—'}",
            f"Forma Rengi        : {self.v_renk.get().strip() or '—'}",
            f"Sezon              : {self.v_sezon.get().strip() or '—'}",
            f"Durum              : {self.v_durum.get().strip() or '—'}",
            f"Aidat Durumu       : {'Ödendi' if self.v_aidat.get() else 'Ödenmedi'}",
            "",
            "Gerekli Evrak Kontrolü:",
            "01. [ ] Kulüp yeniden kayıt / aktif üyelik başvuru formu",
            "02. [ ] Kulüp yetkili ve iletişim bilgileri",
            "03. [ ] Aidat ödeme / üyelik durumu teyidi",
            "",
            "Kulüp Yetkilisi İmza : ____________________",
            "Federasyon Onay      : ____________________",
        ]
        return "\n".join(satirlar)

    def _kulup_form_onizleme_ac(self):
        try:
            metin = self._kulup_form_metni_olustur()
        except ValueError as exc:
            messagebox.showwarning("Uyarı", str(exc))
            return

        wnd = tk.Toplevel(self)
        wnd.title("Kulüp Kayıt Formu Önizleme")
        wnd.configure(bg=BG)
        wnd.geometry("860x680")

        txt = tk.Text(wnd, wrap="word", font=("Consolas", 11))
        txt.pack(fill="both", expand=True, padx=8, pady=(8, 4))
        txt.insert("1.0", metin)

        alt = ttk.Frame(wnd)
        alt.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(alt, text="💾 TXT Olarak Kaydet", style="Neu.TButton",
                   command=lambda: self._kulup_formu_kaydet(txt.get("1.0", "end-1c"), sor=True)).pack(side="left", padx=4)
        ttk.Button(alt, text="🖨 Yazdır", style="Neu.TButton",
                   command=lambda: self._kulup_formu_yazdir(txt.get("1.0", "end-1c"))).pack(side="left", padx=4)
        ttk.Button(alt, text="Kapat", style="Neu.TButton",
                   command=wnd.destroy).pack(side="right", padx=4)

    def _kulup_formu_kaydet(self, icerik: str, sor: bool = False) -> str:
        if sor:
            from tkinter import filedialog
            yol = filedialog.asksaveasfilename(
                title="Kulüp formunu kaydet",
                defaultextension=".txt",
                filetypes=[("Metin Dosyası", "*.txt"), ("Tüm Dosyalar", "*.*")],
                initialfile="kulup_yeniden_kayit_formu.txt",
            )
            if not yol:
                return ""
        else:
            fd, yol = tempfile.mkstemp(prefix="ktbf_kulup_form_", suffix=".txt")
            os.close(fd)
        with open(yol, "w", encoding="utf-8") as dosya:
            dosya.write(icerik)
        return yol

    def _kulup_formu_yazdir(self, icerik: str | None = None):
        try:
            metin = icerik if icerik is not None else self._kulup_form_metni_olustur()
        except ValueError as exc:
            messagebox.showwarning("Uyarı", str(exc))
            return

        try:
            yol = self._kulup_formu_kaydet(metin, sor=False)
            if not yol:
                return
            if os.name == "nt":
                os.startfile(yol, "print")
                messagebox.showinfo("Yazdırma", "Kulüp kayıt formu yazdırma komutu gönderildi.")
            else:
                messagebox.showinfo("Bilgi", f"Yazdırma yalnızca Windows'ta otomatik desteklenir.\nDosya: {yol}")
        except Exception as exc:
            messagebox.showerror("Hata", f"Kulüp kayıt formu yazdırılamadı: {exc}")


# ---------------------------------------------------------------------------
# SPORCU sekmesi
# ---------------------------------------------------------------------------

class SporcuSekme(ttk.Frame):
    _TUM_KATEGORILER = ["U13", "U15", "U17", "U19", "Elite",
                        "Master A", "Master B", "Master C"]

    _FILTRE_KATEGORILER = ["Tümü", "U13", "U15", "U17", "U19", "Elite",
                          "Master A", "Master B", "Master C", "Kategori Dışı"]

    def __init__(self, parent):
        super().__init__(parent, style="TFrame")
        self._secili_id = None
        self._kulup_map: dict = {}   # kulüp adı → id  (None = Ferdi)
        self._grup_modu = False      # gruplandırma aktif mi
        self._build()
        self._yukle_kulupler()
        self._yukle_kategori_filtresi()
        self._listele()

    # ------------------------------------------------------------------
    def _yukle_kulupler(self):
        """Aktif kulüpleri combobox'a yükler."""
        rows = db.kulupler_dropdown()
        self._kulup_map = {"— Ferdi —": None}
        self._kulup_map.update({r["ad"]: r["id"] for r in rows})
        self.cb_kulup["values"] = list(self._kulup_map.keys())
        # Kulüp filtresi combobox'ını da güncelle
        if hasattr(self, 'cb_kulup_filtre'):
            kulup_adlari = ["Tümü"] + [r["ad"] for r in rows]
            self.cb_kulup_filtre["values"] = kulup_adlari

    def _yukle_kategori_filtresi(self):
        """Kategori filtresi combobox'ını yükler."""
        self.cb_kat_filtre["values"] = self._FILTRE_KATEGORILER

    # ------------------------------------------------------------------
    def _build(self):
        ttk.Label(self, text="SPORCU KAYIT VE YÖNETİMİ",
                  style="Header.TLabel").pack(fill="x")

        ust_buton_seridi = ttk.Frame(self)
        ust_buton_seridi.pack(fill="x", padx=_scaled(8), pady=(_scaled(6), 0))
        ttk.Button(ust_buton_seridi, text="➕ Kaydet", style="Add.TButton",
                   command=self._kaydet).pack(side="left", padx=4)
        ttk.Button(ust_buton_seridi, text="✏️ Güncelle", style="Upd.TButton",
                   command=self._guncelle).pack(side="left", padx=4)
        ttk.Button(ust_buton_seridi, text="🗑 Sil", style="Del.TButton",
                   command=self._sil).pack(side="left", padx=4)
        ttk.Button(ust_buton_seridi, text="🔄 Kulüpleri Yenile", style="Neu.TButton",
                   command=self._yukle_kulupler).pack(side="left", padx=4)
        ttk.Button(ust_buton_seridi, text="🌐 Başka Fed. Lisansı", style="Neu.TButton",
               command=self._baska_fed_lisansi_ac).pack(side="left", padx=4)
        ttk.Button(ust_buton_seridi, text="✖ Temizle", style="Neu.TButton",
                   command=self._temizle).pack(side="left", padx=4)

        ttk.Label(ust_buton_seridi, text="| Evrak:").pack(side="left", padx=(10, 2))
        self.v_evrak_saglik = tk.IntVar(value=0)
        self.v_evrak_veli = tk.IntVar(value=0)
        self.v_evrak_baska_fed = tk.IntVar(value=0)
        ttk.Checkbutton(ust_buton_seridi, text="Sağlık Raporu",
                variable=self.v_evrak_saglik).pack(side="left", padx=2)
        ttk.Checkbutton(ust_buton_seridi, text="Veli Muvafakati (<18)",
                variable=self.v_evrak_veli).pack(side="left", padx=2)
        ttk.Checkbutton(ust_buton_seridi, text="Başka Federasyon Beyanı",
                variable=self.v_evrak_baska_fed).pack(side="left", padx=2)

        pane = ttk.PanedWindow(self, orient="horizontal")
        pane.pack(fill="both", expand=True, padx=8, pady=8)

        # Sol: form
        form_frame = ttk.LabelFrame(pane, text="Sporcu Bilgileri", padding=8)
        pane.add(form_frame, weight=1)

        self.v_ad       = _lbl_entry(form_frame, "Ad *",           0)
        self.v_soyad    = _lbl_entry(form_frame, "Soyad *",        1)
        self.v_kimlik   = _lbl_entry(form_frame, "Kimlik No *",    2)
        self.v_dogum    = _lbl_entry(form_frame, "Doğum Tarihi",   3, width=14)
        ttk.Label(form_frame, text="(YYYY-AA-GG)",
                  font=FONT_S, foreground="gray"
                  ).grid(row=3, column=2, sticky="w")
        self.v_cinsiyet = _lbl_combo(form_frame, "Cinsiyet",
                         ["Belirtilmedi", "Erkek", "Kadın"], 4, width=14)
        self.v_cinsiyet.set("Belirtilmedi")
        self.v_uyruk    = _lbl_combo(form_frame, "Uyruk",
                         ["KKTC", "TC", "Diğer"], 5, width=10)
        self.v_uyruk.set("KKTC")
        self.v_pasaport = _lbl_entry(form_frame, "Pasaport No",    6)
        self.v_tel      = _lbl_entry(form_frame, "Telefon",        7)
        self.v_email    = _lbl_entry(form_frame, "E-posta",        8)
        self.v_adres    = _lbl_entry(form_frame, "Adres",          9)
        self.v_sd_kayit = _lbl_check(form_frame,
                         "Spor Dairesi BYS Kayıtlı", 10)

        # Kulüp seçimi — kayıtlı kulüpler veya Ferdi
        ttk.Label(form_frame, text="Kulüp").grid(
            row=11, column=0, sticky="e", padx=(8, 4), pady=4)
        self.v_kulup = tk.StringVar(value="— Ferdi —")
        self.cb_kulup = ttk.Combobox(form_frame, textvariable=self.v_kulup,
                                     width=14, state="readonly")
        self.cb_kulup.grid(row=11, column=1,
                           sticky="w", padx=(0, 8), pady=4)

        # Lisans türü
        self.v_lisans_turu = _lbl_combo(
            form_frame, "Lisans Türü",
            ["Ulusal", "Uluslararası", "Ferdi", "Geçici", "MisafirSporcu"],
            12, width=14)
        self.v_lisans_turu.set("Ulusal")

        # Sezon
        self.v_sezon = _lbl_entry(form_frame, "Sezon", 13, width=10)
        self.v_sezon.set("2026")

        # Üretilen lisans no (salt okunur, kayıt sonrası dolar)
        ttk.Label(form_frame, text="Lisans No").grid(
            row=14, column=0, sticky="e", padx=(8, 4), pady=4)
        self.v_lisans_no = tk.StringVar(value="—")
        ttk.Label(form_frame, textvariable=self.v_lisans_no,
                  foreground=ACCENT, font=FONT_B).grid(
            row=14, column=1, sticky="w", padx=(0, 8), pady=4)

        ttk.Label(form_frame, text="Yaş Kategorisi").grid(
            row=15, column=0, sticky="e", padx=(8, 4), pady=4)
        self.v_yas_kategorisi = tk.StringVar(value="—")
        ttk.Label(form_frame, textvariable=self.v_yas_kategorisi,
                  font=FONT_B).grid(
            row=15, column=1, sticky="w", padx=(0, 8), pady=4)

        ttk.Label(form_frame, text="Yarış Kategorisi").grid(
            row=16, column=0, sticky="e", padx=(8, 4), pady=4)
        self.v_yaris_kategorisi = tk.StringVar(value="—")
        self.cb_yaris_kategorisi = ttk.Combobox(
            form_frame,
            textvariable=self.v_yaris_kategorisi,
            values=self._TUM_KATEGORILER,
            width=14,
            state="readonly",
        )
        self.cb_yaris_kategorisi.grid(
            row=16, column=1, sticky="w", padx=(0, 8), pady=4)
        self.cb_yaris_kategorisi.bind("<<ComboboxSelected>>",
                                       self._yaris_kategorisi_degisti)

        ttk.Label(form_frame, text="MTB Kategorisi").grid(
            row=17, column=0, sticky="e", padx=(8, 4), pady=4)
        self.v_mtb_kategorisi = tk.StringVar(value="—")
        self.cb_mtb_kategorisi = ttk.Combobox(
            form_frame,
            textvariable=self.v_mtb_kategorisi,
            values=self._TUM_KATEGORILER,
            width=14,
            state="readonly",
        )
        self.cb_mtb_kategorisi.grid(
            row=17, column=1, sticky="w", padx=(0, 8), pady=4)
        self.cb_mtb_kategorisi.bind("<<ComboboxSelected>>",
                                     self._mtb_kategorisi_degisti)

        form_frame.columnconfigure(1, weight=1)

        # Sağ: liste + arama + filtre
        list_frame = ttk.LabelFrame(pane, text="Sporcu Listesi", padding=4)
        pane.add(list_frame, weight=2)

        # Arama satırı
        srch_frame = ttk.Frame(list_frame)
        srch_frame.pack(fill="x", pady=(0, 2))
        ttk.Label(srch_frame, text="Ad / Kimlik No:").pack(side="left", padx=4)
        self.v_arama = tk.StringVar()
        ttk.Entry(srch_frame, textvariable=self.v_arama, width=20
                  ).pack(side="left", padx=2)
        ttk.Button(srch_frame, text="Ara", style="Neu.TButton",
                   command=self._ara).pack(side="left", padx=4)
        ttk.Button(srch_frame, text="Tümü", command=self._listele
                   ).pack(side="left")

        # Filtre satırı
        filtre_frame = ttk.Frame(list_frame)
        filtre_frame.pack(fill="x", pady=(0, 2))
        ttk.Label(filtre_frame, text="Kategori:").pack(side="left", padx=(4, 2))
        self.v_kat_filtre = tk.StringVar(value="Tümü")
        self.cb_kat_filtre = ttk.Combobox(filtre_frame, textvariable=self.v_kat_filtre,
                                          values=self._FILTRE_KATEGORILER,
                                          width=12, state="readonly")
        self.cb_kat_filtre.pack(side="left", padx=2)
        self.cb_kat_filtre.bind("<<ComboboxSelected>>", self._filtre_uygula)

        ttk.Label(filtre_frame, text="Kulüp:").pack(side="left", padx=(8, 2))
        self.v_kulup_filtre = tk.StringVar(value="Tümü")
        self.cb_kulup_filtre = ttk.Combobox(filtre_frame, textvariable=self.v_kulup_filtre,
                                            values=["Tümü"],
                                            width=18, state="readonly")
        self.cb_kulup_filtre.pack(side="left", padx=2)
        self.cb_kulup_filtre.bind("<<ComboboxSelected>>", self._filtre_uygula)

        ttk.Separator(filtre_frame, orient="vertical").pack(side="left", fill="y", padx=6)
        self.v_grup_modu = tk.BooleanVar(value=False)
        self.btn_grup = ttk.Checkbutton(filtre_frame, text="📂 Grupla",
                                        variable=self.v_grup_modu,
                                        command=self._grup_modu_degisti)
        self.btn_grup.pack(side="left", padx=4)
        ttk.Button(filtre_frame, text="🔄 Filtreyi Temizle", command=self._filtre_temizle
                   ).pack(side="left", padx=4)
        ttk.Button(filtre_frame, text="📊 Excel Oluştur", style="Neu.TButton",
                   command=self._sporculari_excel_export).pack(side="left", padx=4)

        cols = [("id","ID",38), ("ad","Ad",85), ("soyad","Soyad",95),
            ("cinsiyet","Cinsiyet",80),
            ("kimlik_no","Kimlik No",105), ("dogum_tarihi","Doğum",82),
            ("yas_kategorisi","Yaş Kategorisi",120),
            ("yaris_kategorisi","Yarış Kategorisi",120),
            ("mtb_kategorisi","MTB Kategorisi",120),
            ("uyruk","Uyruk",48), ("telefon","Telefon",100),
            ("lisans_no","Lisans No",88), ("kulup_adi","Kulüp",130),
            ("spor_dairesi_kayitli","BYS",38)]
        tf, self.tree = _make_tree(list_frame, cols)
        tf.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_sec)

    @staticmethod
    def _hesapla_yas(dogum_tarihi: str):
        if not dogum_tarihi:
            return None
        try:
            return date.today().year - int(dogum_tarihi[:4])
        except Exception:
            return None

    @staticmethod
    def _hesapla_yas_kategorisi(dogum_tarihi: str) -> str:
        if not dogum_tarihi:
            return "—"
        try:
            yas = date.today().year - int(dogum_tarihi[:4])
        except Exception:
            return "—"
        if 11 <= yas <= 12:
            return "U13"
        if 13 <= yas <= 14:
            return "U15"
        if 15 <= yas <= 16:
            return "U17"
        if 17 <= yas <= 18:
            return "U19"
        if 19 <= yas <= 34:
            return "Elite"
        if 35 <= yas <= 39:
            return "Master A"
        if 40 <= yas <= 44:
            return "Master B"
        if yas >= 45:
            return "Master C"
        return "Kategori Dışı"

    def _kategori_gorunumu_guncelle(self):
        yas_kat = self._hesapla_yas_kategorisi(self.v_dogum.get().strip())
        self.v_yas_kategorisi.set(yas_kat)

        # Combobox seçeneklerini yaş kategorisine göre belirle
        # Masterlar: kendi kategorisi + bir alt master + Elite
        # Master C → [C, B, A, Elite];  Master B → [B, A, Elite];  Master A → [A, Elite]
        _MASTER_SECENEK = {
            "Master A": ["Master A", "Elite"],
            "Master B": ["Master B", "Master A", "Elite"],
            "Master C": ["Master C", "Master B", "Master A", "Elite"],
        }
        if yas_kat in _MASTER_SECENEK:
            secenekler = _MASTER_SECENEK[yas_kat]
        elif yas_kat in ("—", "Kategori Dışı", "KD"):
            secenekler = self._TUM_KATEGORILER
        else:
            secenekler = [yas_kat]
        self.cb_yaris_kategorisi["values"] = secenekler
        self.cb_mtb_kategorisi["values"] = secenekler

        sezon = self.v_sezon.get().strip() or "2026"
        yaris_kat = yas_kat
        mtb_kat = yas_kat
        if self._secili_id and sezon:
            kayit = db.sporcu_sezon_kategori_getir(self._secili_id, sezon)
            if kayit:
                yaris_kat = kayit["yaris_kategorisi"] or yas_kat
                mtb_kat = kayit["mtb_kategorisi"] or yas_kat
            else:
                secim = db.sporcu_sezon_kayitli_kategori(self._secili_id, sezon)
                if secim:
                    yaris_kat = secim

        self.v_yaris_kategorisi.set(yaris_kat)
        self.v_mtb_kategorisi.set(mtb_kat)
        # Combobox yazı rengi: override varsa mavi, yoksa normal
        self.cb_yaris_kategorisi.configure(
            foreground=ACCENT if yaris_kat != yas_kat else "black"
        )
        self.cb_mtb_kategorisi.configure(
            foreground=ACCENT if mtb_kat != yas_kat else "black"
        )

    def _yaris_kategorisi_degisti(self, _=None):
        """Combobox'tan seçilen yol yarış kategorisini doğrudan kaydeder."""
        self._kategorileri_kaydet()

    def _mtb_kategorisi_degisti(self, _=None):
        """Combobox'tan seçilen MTB kategorisini doğrudan kaydeder."""
        self._kategorileri_kaydet()

    def _kategorileri_kaydet(self):
        if not self._secili_id:
            return
        yaris_kat = self.v_yaris_kategorisi.get().strip()
        mtb_kat = self.v_mtb_kategorisi.get().strip()
        if not yaris_kat or yaris_kat == "—" or not mtb_kat or mtb_kat == "—":
            return
        yas_kat = self._hesapla_yas_kategorisi(self.v_dogum.get().strip())
        sezon = self.v_sezon.get().strip() or "2026"

        if yaris_kat == yas_kat and mtb_kat == yas_kat:
            db.sporcu_sezon_kategori_sil(self._secili_id, sezon)
        else:
            db.sporcu_sezon_kategori_ata(
                self._secili_id,
                sezon,
                yas_kategorisi=yas_kat,
                yaris_kategorisi=yaris_kat,
                mtb_kategorisi=mtb_kat,
            )
        self._kategori_gorunumu_guncelle()

    def _evrak_kontrolu_gecerli(self) -> bool:
        yas = self._hesapla_yas(self.v_dogum.get().strip())
        if yas is not None and yas < 18 and not self.v_evrak_veli.get():
            messagebox.showwarning(
                "Uyarı",
                "18 yaş altı sporcu için Veli Muvafakati zorunludur.",
            )
            return False
        return True

    # ------------------------------------------------------------------
    _SPORCU_QUERY = """
        SELECT s.id, s.ad, s.soyad, s.cinsiyet, s.kimlik_no, s.dogum_tarihi,
               CASE
                   WHEN s.dogum_tarihi IS NULL OR TRIM(s.dogum_tarihi) = '' THEN '—'
                   WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 11 AND 12
                       THEN 'U13'
                   WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 13 AND 14
                       THEN 'Yıldız B / U15'
                   WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 15 AND 16
                       THEN 'Yıldız A / U17'
                   WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 17 AND 18
                       THEN 'Genç / U19'
                   WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 19 AND 34
                       THEN 'Elite'
                   WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 35 AND 39
                       THEN 'Master A'
                   WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 40 AND 44
                       THEN 'Master B'
                   WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) >= 45
                       THEN 'Master C'
                   ELSE 'Kategori Dışı'
               END AS yas_kategorisi,
               COALESCE(
                   ssk.yaris_kategorisi,
                   CASE
                       WHEN s.dogum_tarihi IS NULL OR TRIM(s.dogum_tarihi) = '' THEN '—'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 11 AND 12
                           THEN 'U13'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 13 AND 14
                           THEN 'U15'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 15 AND 16
                           THEN 'U17'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 17 AND 18
                           THEN 'U19'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 19 AND 34
                           THEN 'Elite'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 35 AND 39
                           THEN 'Master A'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 40 AND 44
                           THEN 'Master B'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) >= 45
                           THEN 'Master C'
                       ELSE 'Kategori Dışı'
                   END
               ) AS yaris_kategorisi,
               COALESCE(
                   ssk.mtb_kategorisi,
                   CASE
                       WHEN s.dogum_tarihi IS NULL OR TRIM(s.dogum_tarihi) = '' THEN '—'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 11 AND 12
                           THEN 'U13'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 13 AND 14
                           THEN 'U15'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 15 AND 16
                           THEN 'U17'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 17 AND 18
                           THEN 'U19'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 19 AND 34
                           THEN 'Elite'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 35 AND 39
                           THEN 'Master A'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 40 AND 44
                           THEN 'Master B'
                       WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) >= 45
                           THEN 'Master C'
                       ELSE 'Kategori Dışı'
                   END
               ) AS mtb_kategorisi,
               s.uyruk, s.telefon,
               COALESCE(l.lisans_no, '—')  AS lisans_no,
               COALESCE(k.ad, 'Ferdi')     AS kulup_adi,
               s.spor_dairesi_kayitli
        FROM sporcular s
        LEFT JOIN lisanslar l ON l.id = (
            SELECT id FROM lisanslar
            WHERE sporcu_id = s.id AND durum = 'Aktif'
            ORDER BY id DESC LIMIT 1
        )
        LEFT JOIN kulupler k ON k.id = l.kulup_id
        LEFT JOIN sporcu_sezon_kategorileri ssk
               ON ssk.sporcu_id = s.id AND ssk.sezon = l.sezon
    """

    def _filtre_kosulu(self) -> tuple[str, list]:
        """Seçili filtrelere göre WHERE koşulu ve parametre listesi üretir."""
        kosullar = []
        params = []

        kat = self.v_kat_filtre.get().strip()
        if kat and kat != "Tümü":
            kosullar.append(
                """CASE
                    WHEN s.dogum_tarihi IS NULL OR TRIM(s.dogum_tarihi) = '' THEN '—'
                    WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 11 AND 12 THEN 'U13'
                    WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 13 AND 14 THEN 'U15'
                    WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 15 AND 16 THEN 'U17'
                    WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 17 AND 18 THEN 'U19'
                    WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 19 AND 34 THEN 'Elite'
                    WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 35 AND 39 THEN 'Master A'
                    WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) BETWEEN 40 AND 44 THEN 'Master B'
                    WHEN (CAST(strftime('%Y','now') AS INTEGER) - CAST(substr(s.dogum_tarihi,1,4) AS INTEGER)) >= 45 THEN 'Master C'
                    ELSE 'Kategori Dışı'
                END = ?"""
            )
            params.append(kat)

        kulup = self.v_kulup_filtre.get().strip()
        if kulup and kulup != "Tümü":
            kosullar.append("COALESCE(k.ad, 'Ferdi') = ?")
            params.append(kulup)

        where = ""
        if kosullar:
            where = " WHERE " + " AND ".join(kosullar)
        return where, params

    def _listele(self, *_):
        where, params = self._filtre_kosulu()
        sql = self._SPORCU_QUERY + where + " ORDER BY s.soyad, s.ad"
        with db.get_conn() as conn:
            rows = conn.execute(sql, params).fetchall()

        if self._grup_modu:
            self._listele_gruplu(rows)
        else:
            _fill_tree(self.tree, rows)

    def _ara(self):
        q = self.v_arama.get().strip()
        if not q:
            self._listele()
            return
        where, params = self._filtre_kosulu()
        ara_kosul = " (s.kimlik_no=? OR (s.ad||' '||s.soyad) LIKE ?)"
        if where:
            where += " AND" + ara_kosul
        else:
            where = " WHERE" + ara_kosul
        params.extend([q, f"%{q}%"])
        sql = self._SPORCU_QUERY + where + " ORDER BY s.soyad, s.ad"
        with db.get_conn() as conn:
            rows = conn.execute(sql, params).fetchall()

        if self._grup_modu:
            self._listele_gruplu(rows)
        else:
            _fill_tree(self.tree, rows)

    def _listele_gruplu(self, rows):
        """Sporcu listesini kategori ve kulübe göre gruplandırarak gösterir."""
        self.tree.delete(*self.tree.get_children())

        # rows'u (yas_kategorisi, kulup_adi) ikilisine göre grupla
        from collections import OrderedDict
        gruplar: dict[tuple[str, str], list] = OrderedDict()
        for r in rows:
            kat = r["yas_kategorisi"] if r["yas_kategorisi"] not in ("—", "") else "Belirsiz"
            kulup = r["kulup_adi"] if r["kulup_adi"] not in ("—", "") else "Ferdi"
            anahtar = (kat, kulup)
            if anahtar not in gruplar:
                gruplar[anahtar] = []
            gruplar[anahtar].append(r)

        # Kategori sıralaması
        KAT_SIRA = {k: i for i, k in enumerate(self._FILTRE_KATEGORILER)}

        def sira_anahtar(g):
            (kat, kulup), _ = g
            return (KAT_SIRA.get(kat, 99), kulup)

        sirali = sorted(gruplar.items(), key=sira_anahtar)

        iid_counter = [0]
        for (kat, kulup), elemanlar in sirali:
            iid_counter[0] += 1
            grup_id = f"grup_{iid_counter[0]}"
            etiket = f"  📁 {kat} — {kulup}  ({len(elemanlar)} sporcu)"
            self.tree.insert("", "end", iid=grup_id, values=[etiket, "", "", "", "", "", "", "", "", "", "", "", "", ""],
                             tags=("group",))

            for i, row in enumerate(elemanlar):
                self.tree.insert(grup_id, "end", iid=str(row[0]),
                                 values=list(row),
                                 tags=("even" if i % 2 == 0 else "odd",))

        self.tree.item(self.tree.get_children()[0]) if self.tree.get_children() else None

    def _filtre_uygula(self, *_):
        """Filtre değiştiğinde listeyi yeniler."""
        if self.v_arama.get().strip():
            self._ara()
        else:
            self._listele()

    def _grup_modu_degisti(self):
        """Gruplandırma butonu değiştiğinde listeyi yeniler."""
        self._grup_modu = self.v_grup_modu.get()
        self._filtre_uygula()

    def _filtre_temizle(self):
        """Tüm filtreleri sıfırlar ve listeyi yeniler."""
        self.v_kat_filtre.set("Tümü")
        self.v_kulup_filtre.set("Tümü")
        self.v_arama.set("")
        self._listele()

    def _sporculari_excel_export(self):
        """Mevcut sporcu listesini kulüp bazında gruplandırarak Excel (.xlsx) olarak dışa aktarır."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from tkinter import filedialog
        from itertools import groupby

        raw_rows = []
        for child in self.tree.get_children():
            # Grup başlıklarını (string ID) atla
            try:
                int(child)
            except ValueError:
                continue
            values = self.tree.item(child, "values")
            if values and len(values) >= 14:
                raw_rows.append(list(values))

        if not raw_rows:
            messagebox.showinfo("Bilgi", "Dışa aktarılacak sporcu bulunamadı.")
            return

        yol = filedialog.asksaveasfilename(
            title="Excel olarak kaydet (Gruplu)",
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx")],
            initialfile="sporcu_listesi_gruplu.xlsx",
        )
        if not yol:
            return

        # Kulüp adına göre sırala (13. sütun = indeks 12)
        raw_rows.sort(key=lambda r: r[12] if r[12] else "")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sporcu Listesi"

            grup_font = Font(bold=True, size=11, color="FFFFFF")
            grup_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
            baslik_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            baslik_font = Font(bold=True, size=11, color="FFFFFF")

            basliklar = ["ID", "Ad", "Soyad", "Cinsiyet", "Kimlik No",
                         "Doğum Tarihi", "Yaş Kategorisi", "Yarış Kategorisi",
                         "MTB Kategorisi", "Uyruk", "Telefon", "Lisans No", "Kulüp", "BYS"]
            KOLON_SAY = len(basliklar)

            # Başlık satırı
            for col, b in enumerate(basliklar, 1):
                h = ws.cell(row=1, column=col, value=b)
                h.font = baslik_font
                h.fill = baslik_fill
                h.alignment = Alignment(horizontal="center")

            satir_no = 2
            for kulup, grp in groupby(raw_rows, key=lambda r: r[12] if r[12] else "Belirtilmemiş"):
                grp_list = list(grp)
                # Grup başlık satırı
                ws.merge_cells(start_row=satir_no, start_column=1,
                               end_row=satir_no, end_column=KOLON_SAY)
                h = ws.cell(row=satir_no, column=1,
                            value=f"📁 {kulup}  ({len(grp_list)} sporcu)")
                h.font = grup_font
                h.fill = grup_fill
                satir_no += 1
                # Sporcu satırları
                for row in grp_list:
                    for j, val in enumerate(row, 1):
                        ws.cell(row=satir_no, column=j, value=val)
                    satir_no += 1

            # Sütun genişlikleri
            genislikler = [6, 18, 18, 12, 18, 14, 18, 18, 18, 8, 16, 16, 22, 6]
            for col, w in enumerate(genislikler, 1):
                ws.column_dimensions[chr(64 + col)].width = w

            wb.save(yol)
            messagebox.showinfo("Başarılı", f"Dışa aktarıldı:\n{yol}")
        except Exception as exc:
            messagebox.showerror("Hata", f"Dışa aktarılamadı: {exc}")

    # ------------------------------------------------------------------
    def _on_sec(self, _=None):
        sel = self.tree.selection()
        if not sel:
            return
        try:
            self._secili_id = int(sel[0])
        except ValueError:
            # Grup başlığı seçilmiş, işlem yapma
            return
        row = db.sporcu_getir(self._secili_id)
        if row:
            self.v_ad.set(row["ad"] or "")
            self.v_soyad.set(row["soyad"] or "")
            self.v_kimlik.set(row["kimlik_no"] or "")
            self.v_dogum.set(row["dogum_tarihi"] or "")
            self.v_cinsiyet.set(row["cinsiyet"] or "Belirtilmedi")
            self.v_uyruk.set(row["uyruk"] or "KKTC")
            self.v_pasaport.set(row["pasaport_no"] or "")
            self.v_tel.set(row["telefon"] or "")
            self.v_email.set(row["email"] or "")
            self.v_adres.set(row["adres"] or "")
            self.v_sd_kayit.set(row["spor_dairesi_kayitli"] or 0)
        # Aktif lisans bilgilerini doldur
        with db.get_conn() as conn:
            lis = conn.execute(
                """SELECT l.lisans_no, l.lisans_turu, l.sezon,
                          l.saglik_raporu, l.veli_muvafakati, l.baska_federasyon_beyani,
                          k.ad AS kulup_adi
                   FROM lisanslar l
                   LEFT JOIN kulupler k ON k.id = l.kulup_id
                   WHERE l.sporcu_id=? AND l.durum='Aktif'
                   ORDER BY l.id DESC LIMIT 1""",
                (self._secili_id,)
            ).fetchone()
        if lis:
            self.v_lisans_no.set(lis["lisans_no"] or "—")
            self.v_lisans_turu.set(lis["lisans_turu"] or "Ulusal")
            self.v_sezon.set(lis["sezon"] or "2026")
            kulup_ad = lis["kulup_adi"] or "— Ferdi —"
            self.v_kulup.set(kulup_ad if kulup_ad in self._kulup_map
                             else "— Ferdi —")
            self.v_evrak_saglik.set(lis["saglik_raporu"] or 0)
            self.v_evrak_veli.set(lis["veli_muvafakati"] or 0)
            self.v_evrak_baska_fed.set(lis["baska_federasyon_beyani"] or 0)
        else:
            self.v_lisans_no.set("—")
            self.v_evrak_saglik.set(0)
            self.v_evrak_veli.set(0)
            self.v_evrak_baska_fed.set(0)
        self._kategori_gorunumu_guncelle()

    # ------------------------------------------------------------------
    def _kaydet(self):
        if not all([self.v_ad.get().strip(),
                    self.v_soyad.get().strip(),
                    self.v_kimlik.get().strip()]):
            messagebox.showwarning("Uyarı", "Ad, Soyad ve Kimlik No zorunludur.")
            return
        if not self._evrak_kontrolu_gecerli():
            return
        try:
            sid = db.sporcu_ekle(
                self.v_ad.get().strip(),
                self.v_soyad.get().strip(),
                self.v_kimlik.get().strip(),
                dogum_tarihi=self.v_dogum.get() or None,
                cinsiyet=self.v_cinsiyet.get() or "Belirtilmedi",
                uyruk=self.v_uyruk.get() or "KKTC",
                pasaport_no=self.v_pasaport.get() or None,
                telefon=self.v_tel.get() or None,
                email=self.v_email.get() or None,
                adres=self.v_adres.get() or None,
                spor_dairesi_kayitli=self.v_sd_kayit.get(),
            )
            kulup_id = self._kulup_map.get(self.v_kulup.get())  # None = Ferdi
            lid = db.lisans_ekle(
                sid,
                self.v_lisans_turu.get() or "Ulusal",
                self.v_sezon.get() or "2026",
                kulup_id=kulup_id,
                saglik_raporu=self.v_evrak_saglik.get(),
                veli_muvafakati=self.v_evrak_veli.get(),
                baska_federasyon_beyani=self.v_evrak_baska_fed.get(),
            )
            with db.get_conn() as conn:
                lis = conn.execute(
                    "SELECT lisans_no FROM lisanslar WHERE id=?", (lid,)
                ).fetchone()
            no = lis["lisans_no"] if lis else "—"
            self.v_lisans_no.set(no)
            self._secili_id = sid
            self._kategori_gorunumu_guncelle()
            messagebox.showinfo("Başarılı",
                                f"Sporcu kaydedildi.\nLisans No: {no}")
            self._listele()
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    # ------------------------------------------------------------------
    def _guncelle(self):
        if not self._secili_id:
            messagebox.showwarning("Uyarı", "Listeden bir sporcu seçin.")
            return
        if not self._evrak_kontrolu_gecerli():
            return
        try:
            kulup_id = self._kulup_map.get(self.v_kulup.get())
            db.sporcu_guncelle(
                self._secili_id,
                ad=self.v_ad.get().strip(),
                soyad=self.v_soyad.get().strip(),
                kimlik_no=self.v_kimlik.get().strip(),
                dogum_tarihi=self.v_dogum.get() or None,
                cinsiyet=self.v_cinsiyet.get() or "Belirtilmedi",
                uyruk=self.v_uyruk.get() or "KKTC",
                pasaport_no=self.v_pasaport.get() or None,
                telefon=self.v_tel.get() or None,
                email=self.v_email.get() or None,
                adres=self.v_adres.get() or None,
                spor_dairesi_kayitli=self.v_sd_kayit.get(),
            )
            db.aktif_lisans_guncelle(
                self._secili_id,
                lisans_turu=self.v_lisans_turu.get() or "Ulusal",
                sezon=self.v_sezon.get() or "2026",
                kulup_id=kulup_id,
                saglik_raporu=self.v_evrak_saglik.get(),
                veli_muvafakati=self.v_evrak_veli.get(),
                baska_federasyon_beyani=self.v_evrak_baska_fed.get(),
            )
            self._kategori_gorunumu_guncelle()
            self._listele()
            messagebox.showinfo("Başarılı", "Sporcu güncellendi.")
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    def _sil(self):
        if not self._secili_id:
            messagebox.showwarning("Uyarı", "Listeden bir sporcu seçin.")
            return
        if not messagebox.askyesno("Onay", "Seçili sporcu silinsin mi?"):
            return
        try:
            db.sporcu_sil(self._secili_id)
            self._temizle()
            self._listele()
            messagebox.showinfo("Başarılı", "Sporcu silindi.")
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    # ------------------------------------------------------------------
    def _temizle(self):
        for v in (self.v_ad, self.v_soyad, self.v_kimlik, self.v_dogum,
                  self.v_pasaport, self.v_tel, self.v_email, self.v_adres):
            v.set("")
        self.v_cinsiyet.set("Belirtilmedi")
        self.v_uyruk.set("KKTC")
        self.v_sd_kayit.set(0)
        self.v_kulup.set("— Ferdi —")
        self.v_lisans_turu.set("Ulusal")
        self.v_sezon.set("2026")
        self.v_lisans_no.set("—")
        self.v_evrak_saglik.set(0)
        self.v_evrak_veli.set(0)
        self.v_evrak_baska_fed.set(0)
        self.v_yas_kategorisi.set("—")
        self.v_yaris_kategorisi.set("—")
        self.v_mtb_kategorisi.set("—")
        self.cb_yaris_kategorisi.configure(foreground="black")
        self.cb_mtb_kategorisi.configure(foreground="black")
        self._secili_id = None
        self.tree.selection_remove(*self.tree.selection())

    def _baska_fed_lisansi_ac(self):
        """Seçili sporcunun yabancı federasyon lisanslarını yönetmek için pencere açar."""
        if not self._secili_id:
            messagebox.showwarning("Uyarı", "Önce listeden bir sporcu seçin.")
            return
        BaskaFederasyonLisansPenceresi(self, self._secili_id)


# ---------------------------------------------------------------------------
# Başka Federasyon Lisansı Yönetim Penceresi
# ---------------------------------------------------------------------------

class BaskaFederasyonLisansPenceresi(tk.Toplevel):
    """Yabancı federasyon lisansı giriş / düzenleme / silme penceresi."""

    def __init__(self, parent: tk.Misc, sporcu_id: int):
        super().__init__(parent)
        self._sporcu_id = sporcu_id
        self._secili_id: int | None = None
        self.title("Başka Federasyon Lisansı Yönetimi")
        self.configure(bg=BG)
        _apply_window_geometry(self, width_ratio=0.60, height_ratio=0.62,
                               min_width=700, min_height=520)

        # Sporcu bilgisini al
        with db.get_conn() as conn:
            row = conn.execute(
                "SELECT ad, soyad, kimlik_no FROM sporcular WHERE id=?",
                (sporcu_id,)
            ).fetchone()
        sporcu_etiket = f"{row['ad']} {row['soyad']} ({row['kimlik_no']})" if row else f"#{sporcu_id}"

        ttk.Label(self, text=f"🌐 Başka Federasyon Lisansı — {sporcu_etiket}",
                  style="Header.TLabel").pack(fill="x")

        # ── Form ──────────────────────────────────────────────────────
        form = ttk.LabelFrame(self, text="Yabancı Federasyon Lisans Bilgisi", padding=8)
        form.pack(fill="x", padx=8, pady=8)
        form.columnconfigure(1, weight=1)

        self.v_federasyon = _lbl_entry(form, "Federasyon *", 0, width=24)
        self.v_kulup_ad    = _lbl_entry(form, "Kulüp Adı",    1, width=24)
        self.v_lisans_no   = _lbl_entry(form, "Lisans No",     2, width=24)
        self.v_gecerlilik  = _lbl_entry(form, "Geçerlilik Tarihi", 3, width=14)
        ttk.Label(form, text="(YYYY-AA-GG)", font=FONT_S,
                  foreground="gray").grid(row=3, column=2, sticky="w")
        self.v_muvafakat   = _lbl_check(form, "Kulüp Muvafakati Var", 4)

        btn_row = ttk.Frame(form)
        btn_row.grid(row=5, column=0, columnspan=3, pady=(10, 2))
        ttk.Button(btn_row, text="➕ Ekle",   style="Add.TButton",
                   command=self._ekle).pack(side="left", padx=4)
        ttk.Button(btn_row, text="✏️ Güncelle", style="Upd.TButton",
                   command=self._guncelle).pack(side="left", padx=4)
        ttk.Button(btn_row, text="🗑 Sil",    style="Del.TButton",
                   command=self._sil).pack(side="left", padx=4)
        ttk.Button(btn_row, text="✖ Temizle", style="Neu.TButton",
                   command=self._temizle).pack(side="left", padx=4)

        # ── Liste ─────────────────────────────────────────────────────
        list_frame = ttk.LabelFrame(self, text="Kayıtlı Lisanslar", padding=4)
        list_frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        cols = [
            ("id",                "ID",          36),
            ("sporcu",            "Sporcu",      200),
            ("yabanci_federasyon", "Federasyon",  160),
            ("kulup",             "Kulüp",       150),
            ("lisans_no",         "Lisans No",   130),
            ("gecerlilik_tarihi", "Geçerlilik",  100),
            ("kulup_muvafakati",  "Muvafakat",    80),
            ("beyan_tarihi",      "Beyan Tarihi",100),
        ]
        tf, self.tree = _make_tree(list_frame, cols)
        tf.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_sec)

        self._listele()

    # ── Yardımcılar ───────────────────────────────────────────────────

    def _listele(self):
        rows = db.yabanci_lisanslari_listele(self._sporcu_id)
        tree_rows = []
        for r in rows:
            tree_rows.append((
                r["id"],
                f"{r['ad']} {r['soyad']}",
                r["yabanci_federasyon"],
                r["kulup"] or "—",
                r["lisans_no"] or "—",
                r["gecerlilik_tarihi"] or "—",
                "✅" if r["kulup_muvafakati"] else "❌",
                r["beyan_tarihi"],
            ))
        self.tree.delete(*self.tree.get_children())
        for i, vals in enumerate(tree_rows):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", iid=str(vals[0]),
                             values=vals, tags=(tag,))

    def _on_sec(self, _=None):
        sel = self.tree.selection()
        if not sel:
            return
        self._secili_id = int(sel[0])
        row = db.yabanci_lisans_getir(self._secili_id)
        if row:
            self.v_federasyon.set(row["yabanci_federasyon"] or "")
            self.v_kulup_ad.set(row["kulup"] or "")
            self.v_lisans_no.set(row["lisans_no"] or "")
            self.v_gecerlilik.set(row["gecerlilik_tarihi"] or "")
            self.v_muvafakat.set(row["kulup_muvafakati"] or 0)

    def _ekle(self):
        fed = self.v_federasyon.get().strip()
        if not fed:
            messagebox.showwarning("Uyarı", "Federasyon adı zorunludur.", parent=self)
            return
        try:
            db.yabanci_lisans_ekle(
                self._sporcu_id,
                yabanci_federasyon=fed,
                kulup=self.v_kulup_ad.get().strip() or None,
                lisans_no=self.v_lisans_no.get().strip() or None,
                gecerlilik_tarihi=self.v_gecerlilik.get().strip() or None,
                kulup_muvafakati=self.v_muvafakat.get(),
            )
            self._temizle()
            self._listele()
            messagebox.showinfo("Başarılı", "Yabancı federasyon lisansı eklendi.", parent=self)
        except Exception as exc:
            messagebox.showerror("Hata", str(exc), parent=self)

    def _guncelle(self):
        if not self._secili_id:
            messagebox.showwarning("Uyarı", "Listeden bir kayıt seçin.", parent=self)
            return
        fed = self.v_federasyon.get().strip()
        if not fed:
            messagebox.showwarning("Uyarı", "Federasyon adı zorunludur.", parent=self)
            return
        try:
            db.yabanci_lisans_guncelle(
                self._secili_id,
                yabanci_federasyon=fed,
                kulup=self.v_kulup_ad.get().strip() or None,
                lisans_no=self.v_lisans_no.get().strip() or None,
                gecerlilik_tarihi=self.v_gecerlilik.get().strip() or None,
                kulup_muvafakati=self.v_muvafakat.get(),
            )
            self._listele()
            messagebox.showinfo("Başarılı", "Yabancı federasyon lisansı güncellendi.", parent=self)
        except Exception as exc:
            messagebox.showerror("Hata", str(exc), parent=self)

    def _sil(self):
        if not self._secili_id:
            messagebox.showwarning("Uyarı", "Listeden bir kayıt seçin.", parent=self)
            return
        if not messagebox.askyesno("Onay", "Bu kayıt silinsin mi?", parent=self):
            return
        try:
            db.yabanci_lisans_sil(self._secili_id)
            self._temizle()
            self._listele()
            messagebox.showinfo("Başarılı", "Yabancı federasyon lisansı silindi.", parent=self)
        except Exception as exc:
            messagebox.showerror("Hata", str(exc), parent=self)

    def _temizle(self):
        self.v_federasyon.set("")
        self.v_kulup_ad.set("")
        self.v_lisans_no.set("")
        self.v_gecerlilik.set("")
        self.v_muvafakat.set(0)
        self._secili_id = None
        self.tree.selection_remove(*self.tree.selection())


# ---------------------------------------------------------------------------
# YARIŞ KAYIT sekmesi
# ---------------------------------------------------------------------------

def _cinsiyet_harf(cinsiyet: str) -> str:
    """Cinsiyet değerini kısa gösterime çevirir: K/E/—."""
    if cinsiyet == "Kadın":
        return "K"
    if cinsiyet == "Erkek":
        return "E"
    return "—"


def _yas_kategorisi_hesapla(dogum_tarihi: str) -> str:
    """Doğum tarihinden yaş kategorisini hesaplar."""
    if not dogum_tarihi:
        return "—"
    try:
        yas = date.today().year - int(dogum_tarihi[:4])
        if 11 <= yas <= 12:  return "U13"
        if 13 <= yas <= 14:  return "U15"
        if 15 <= yas <= 16:  return "U17"
        if 17 <= yas <= 18:  return "U19"
        if 19 <= yas <= 34:  return "Elite"
        if 35 <= yas <= 39:  return "Master A"
        if 40 <= yas <= 44:  return "Master B"
        if yas >= 45:        return "Master C"
        return "KD"
    except Exception:
        return "—"


class HibSekme(ttk.Frame):
    """Herkes İçin Bisiklet sporcularının kayıt ve yönetim ekranı."""

    def __init__(self, parent):
        super().__init__(parent, style="TFrame")
        self._secili_id = None
        self._build()
        self._listele()

    def _build(self):
        ttk.Label(self, text="HİB - HERKES İÇİN BİSİKLET",
                  style="Header.TLabel").pack(fill="x")

        pane = ttk.PanedWindow(self, orient="horizontal")
        pane.pack(fill="both", expand=True, padx=8, pady=8)

        form = ttk.LabelFrame(pane, text="HİB Sporcu Bilgileri", padding=8)
        pane.add(form, weight=1)
        self.v_ad = _lbl_entry(form, "Ad *", 0)
        self.v_soyad = _lbl_entry(form, "Soyad *", 1)
        self.v_dogum = _lbl_entry(form, "Doğum Tarihi *", 2, width=14)
        ttk.Label(form, text="(YYYY-AA-GG)", font=FONT_S, foreground="gray").grid(
            row=2, column=2, sticky="w")
        self.v_kimlik = _lbl_entry(form, "Kimlik No *", 3)
        self.v_cinsiyet = _lbl_combo(
            form, "Cinsiyet", ["Belirtilmedi", "Erkek", "Kadın"], 4, width=14)
        self.v_cinsiyet.set("Belirtilmedi")

        ttk.Label(form, text="Yaş Kategorisi").grid(
            row=5, column=0, sticky="e", padx=(8, 4), pady=4)
        self.v_yas_kategori = tk.StringVar(value="—")
        ttk.Label(form, textvariable=self.v_yas_kategori, font=FONT_B).grid(
            row=5, column=1, sticky="w", padx=(0, 8), pady=4)

        ttk.Label(form, text="Yarış Kategorisi").grid(
            row=6, column=0, sticky="e", padx=(8, 4), pady=4)
        self.v_yaris_kategori = tk.StringVar(value="—")
        ttk.Label(form, textvariable=self.v_yaris_kategori, font=FONT_B).grid(
            row=6, column=1, sticky="w", padx=(0, 8), pady=4)

        self.v_dogum.trace_add("write", self._kategori_guncelle)
        form.columnconfigure(1, weight=1)
        buttons = ttk.Frame(form)
        buttons.grid(row=7, column=0, columnspan=2, pady=(12, 4))
        ttk.Button(buttons, text="➕ Kaydet", style="Add.TButton",
                   command=self._kaydet).pack(side="left", padx=4)
        ttk.Button(buttons, text="✏️ Güncelle", style="Upd.TButton",
                   command=self._guncelle).pack(side="left", padx=4)
        ttk.Button(buttons, text="🗑 Sil", style="Del.TButton",
                   command=self._sil).pack(side="left", padx=4)
        ttk.Button(buttons, text="✖ Temizle", style="Neu.TButton",
                   command=self._temizle).pack(side="left", padx=4)

        liste = ttk.LabelFrame(pane, text="HİB Sporcu Listesi", padding=4)
        pane.add(liste, weight=2)
        cols = [("id", "ID", 42), ("ad", "Ad", 120), ("soyad", "Soyad", 130),
            ("cinsiyet", "Cinsiyet", 85), ("dogum", "Doğum Tarihi", 105),
            ("kimlik", "Kimlik No", 120),
                ("yas", "Yaş Kategorisi", 115), ("yaris", "Yarış Kategorisi", 115),
                ("lisans", "Lisans No", 120)]
        frame, self.tree = _make_tree(liste, cols)
        frame.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_sec)

    def _kategori_guncelle(self, *_):
        kategori = db.hib_kategori_hesapla(self.v_dogum.get().strip())
        self.v_yas_kategori.set(kategori)
        self.v_yaris_kategori.set(kategori)

    def _listele(self):
        rows = []
        for row in db.hib_sporcular_listele():
            kategori = db.hib_kategori_hesapla(row["dogum_tarihi"])
            rows.append((row["id"], row["ad"], row["soyad"], row["cinsiyet"],
                         row["dogum_tarihi"], row["kimlik_no"], kategori, kategori,
                         row["lisans_no"] or "—"))
        _fill_tree(self.tree, rows)

    def _form_gecerli_mi(self) -> bool:
        if not all((self.v_ad.get().strip(), self.v_soyad.get().strip(),
                    self.v_dogum.get().strip(), self.v_kimlik.get().strip())):
            messagebox.showwarning("Uyarı", "Ad, soyad, doğum tarihi ve kimlik no zorunludur.")
            return False
        if self.v_yas_kategori.get() in ("—", "Kategori Dışı"):
            messagebox.showwarning(
                "Uyarı", "HİB kaydı için geçerli bir doğum tarihi ve en az 25 yaş gereklidir.")
            return False
        return True

    def _kaydet(self):
        if not self._form_gecerli_mi():
            return
        try:
            db.hib_sporcu_ekle(self.v_ad.get().strip(), self.v_soyad.get().strip(),
                                self.v_dogum.get().strip(), self.v_kimlik.get().strip(),
                                self.v_cinsiyet.get())
            self._temizle()
            self._listele()
            messagebox.showinfo("Başarılı", "HİB sporcusu kaydedildi.")
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    def _guncelle(self):
        if not self._secili_id:
            messagebox.showwarning("Uyarı", "Listeden bir HİB sporcusu seçin.")
            return
        if not self._form_gecerli_mi():
            return
        try:
            db.sporcu_guncelle(
                self._secili_id, ad=self.v_ad.get().strip(), soyad=self.v_soyad.get().strip(),
                dogum_tarihi=self.v_dogum.get().strip(), kimlik_no=self.v_kimlik.get().strip(),
                cinsiyet=self.v_cinsiyet.get(), hib_sporcusu=1,
            )
            self._listele()
            messagebox.showinfo("Başarılı", "HİB sporcusu güncellendi.")
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    def _sil(self):
        if not self._secili_id:
            messagebox.showwarning("Uyarı", "Listeden bir HİB sporcusu seçin.")
            return
        if not messagebox.askyesno("Onay", "Seçili HİB sporcusu silinsin mi?"):
            return
        try:
            db.hib_sporcu_sil(self._secili_id)
            self._temizle()
            self._listele()
            messagebox.showinfo("Başarılı", "HİB sporcusu silindi.")
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    def _on_sec(self, _=None):
        secim = self.tree.selection()
        if not secim:
            return
        self._secili_id = int(secim[0])
        row = db.sporcu_getir(self._secili_id)
        if row:
            self.v_ad.set(row["ad"] or "")
            self.v_soyad.set(row["soyad"] or "")
            self.v_dogum.set(row["dogum_tarihi"] or "")
            self.v_kimlik.set(row["kimlik_no"] or "")
            self.v_cinsiyet.set(row["cinsiyet"] or "Belirtilmedi")

    def _temizle(self):
        self._secili_id = None
        for value in (self.v_ad, self.v_soyad, self.v_dogum, self.v_kimlik):
            value.set("")
        self.v_cinsiyet.set("Belirtilmedi")
        self.tree.selection_remove(*self.tree.selection())


class YarisKayitSekme(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, style="TFrame")
        self._secili_yaris_id = None
        self._yaris_map: dict = {}
        self._sporcu_map: dict = {}
        self._build()
        self._yarislari_yukle()
        self._sporculari_yukle()
        self._kayitlari_listele()

    def _build(self):
        ttk.Label(self, text="YARIŞ KAYIT MODÜLÜ",
                  style="Header.TLabel").pack(fill="x")

        ust = ttk.LabelFrame(self, text="Yarış Tanımı", padding=8)
        ust.pack(fill="x", padx=8, pady=(8, 4))

        self.v_yaris_ad = _lbl_entry(ust, "Yarış Adı *", 0, width=26)
        self.v_yaris_tarih = _lbl_entry(ust, "Tarih", 1, width=14)
        ttk.Label(ust, text="(YYYY-AA-GG)", font=FONT_S,
                  foreground="gray").grid(row=1, column=2, sticky="w")
        self.v_yaris_yer = _lbl_entry(ust, "Yer", 2, width=20)
        self.v_yaris_disiplin = _lbl_combo(
            ust, "Disiplin",
            ["Yol", "MTB", "Pist", "BMX", "Cyclocross", "Diğer"],
            3, width=14)
        self.v_yaris_disiplin.set("Yol")
        self.v_yaris_sezon = _lbl_entry(ust, "Sezon *", 4, width=10)
        self.v_yaris_sezon.set("2026")
        self.v_yaris_durum = _lbl_combo(
            ust, "Durum",
            ["Planlandı", "Kayıt Açık", "Tamamlandı", "İptal"],
            5, width=14)
        self.v_yaris_durum.set("Kayıt Açık")

        ust.columnconfigure(1, weight=1)

        b1 = ttk.Frame(ust)
        b1.grid(row=6, column=0, columnspan=3, pady=(8, 2))
        ttk.Button(b1, text="➕ Yarış Ekle", style="Add.TButton",
                   command=self._yaris_ekle).pack(side="left", padx=4)
        ttk.Button(b1, text="✏️ Yarış Güncelle", style="Upd.TButton",
               command=self._yaris_guncelle).pack(side="left", padx=4)
        ttk.Button(b1, text="🗑 Yarış Sil", style="Del.TButton",
               command=self._yaris_sil).pack(side="left", padx=4)
        ttk.Button(b1, text="🔄 Yarışları Yenile", style="Neu.TButton",
                   command=self._yarislari_yukle).pack(side="left", padx=4)

        orta = ttk.LabelFrame(self, text="Yarış Listesi", padding=4)
        orta.pack(fill="x", padx=8, pady=4)
        cols_yaris = [
            ("id", "ID", 42),
            ("ad", "Yarış", 220),
            ("tarih", "Tarih", 90),
            ("yer", "Yer", 140),
            ("disiplin", "Disiplin", 90),
            ("sezon", "Sezon", 65),
            ("durum", "Durum", 100),
        ]
        tf1, self.tree_yaris = _make_tree(orta, cols_yaris)
        tf1.pack(fill="x", expand=True)
        self.tree_yaris.bind("<<TreeviewSelect>>", self._on_yaris_sec)
        self.tree_yaris.bind("<Double-1>", self._on_yaris_cift_tikla)

        aksiyon = ttk.Frame(self)
        aksiyon.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Button(aksiyon, text="📝 Sporcu Kayıt Penceresini Aç",
               style="Neu.TButton",
               command=self._kayit_penceresi_ac).pack(side="left", padx=2)

        alt = ttk.LabelFrame(self, text="Sporcu Yarış Kaydı", padding=8)
        # Bu panel artık ana ekranda gösterilmiyor; kayıt işlemi ayrı pencerede.

        ttk.Label(alt, text="Yarış").grid(row=0, column=0, sticky="e",
                                            padx=(8, 4), pady=4)
        self.v_kayit_yaris = tk.StringVar()
        self.cb_kayit_yaris = ttk.Combobox(
            alt, textvariable=self.v_kayit_yaris, width=32, state="readonly")
        self.cb_kayit_yaris.grid(row=0, column=1, sticky="ew", padx=(0, 8), pady=4)
        self.cb_kayit_yaris.bind("<<ComboboxSelected>>", self._on_kayit_yaris_sec)

        ttk.Label(alt, text="Sporcu Ara").grid(row=1, column=0, sticky="e",
                                                padx=(8, 4), pady=4)
        self.v_sporcu_ara = tk.StringVar()
        ttk.Entry(alt, textvariable=self.v_sporcu_ara, width=22
                  ).grid(row=1, column=1, sticky="ew", padx=(0, 4), pady=4)
        ttk.Label(alt, text="→").grid(row=1, column=2, sticky="w", padx=(0, 2), pady=4)
        self.v_kayit_sporcu = tk.StringVar()
        self.cb_kayit_sporcu = ttk.Combobox(
            alt, textvariable=self.v_kayit_sporcu, width=34, state="readonly")
        self.cb_kayit_sporcu.grid(row=1, column=3, sticky="ew", padx=(0, 8), pady=4)

        ttk.Label(alt, text="Kategori").grid(row=2, column=0, sticky="e",
                                               padx=(8, 4), pady=4)
        self.v_kat_otomatik = tk.StringVar(value="—")
        self.lbl_kat = ttk.Label(alt, textvariable=self.v_kat_otomatik, font=FONT_B)
        self.lbl_kat.grid(row=2, column=1, sticky="w", padx=(0, 8), pady=4)

        # Arama yazıldıkça filtrele, sporcu seçilince kategoriyi göster
        self.v_sporcu_ara.trace_add("write", self._sporcu_ara_filtrele)
        self.cb_kayit_sporcu.bind("<<ComboboxSelected>>", self._kategori_goruntule)

        alt.columnconfigure(3, weight=1)

        b2 = ttk.Frame(alt)
        b2.grid(row=3, column=0, columnspan=4, pady=(8, 4))
        ttk.Button(b2, text="➕ Kaydı Oluştur", style="Add.TButton",
                   command=self._kayit_ekle).pack(side="left", padx=4)
        ttk.Button(b2, text="✖ Seçili Kaydı Sil", style="Del.TButton",
                   command=self._kayit_sil).pack(side="left", padx=4)
        ttk.Button(b2, text=" Sporcuları Yenile", style="Neu.TButton",
                   command=self._sporculari_yukle).pack(side="left", padx=4)

        cols_kayit = [
            ("id", "ID", 42),
            ("yaris_adi", "Yarış", 200),
            ("yaris_tarihi", "Yarış Tarihi", 95),
            ("sporcu", "Sporcu", 180),
            ("cinsiyet", "Cinsiyet", 60),
            ("lisans_no", "Lisans No", 90),
            ("kategori", "Kategori", 120),
            ("durum", "Durum", 85),
            ("kayit_tarihi", "Kayıt Tarihi", 95),
        ]
        tf2, self.tree_kayit = _make_tree(alt, cols_kayit)
        tf2.grid(row=4, column=0, columnspan=2, sticky="nsew", pady=(4, 0))
        alt.rowconfigure(4, weight=1)

        # Sağ tık menüsü
        self._kayit_menu = tk.Menu(self.tree_kayit, tearoff=0)
        self._kayit_menu.add_command(label="✏️ Seçili Kaydı Güncelle",
                                     command=self._kayit_guncelle_dialog)
        self._kayit_menu.add_separator()
        self._kayit_menu.add_command(label="📊 Excel'e Aktar (CSV)",
                                     command=self._kayitlari_csv_export)
        self.tree_kayit.bind("<Button-3>", self._kayit_menu_goster)
        self.tree_kayit.bind("<Double-1>", self._on_kayit_cift_tikla)

    def _yaris_label(self, row):
        tarih = row["tarih"] or "Tarihsiz"
        return f"{row['ad']} ({tarih})"

    def _yarislari_yukle(self):
        rows = db.yarislar_listele()
        _fill_tree(self.tree_yaris, rows)
        acik_rows = db.kayit_acik_yarislar()
        self._yaris_map = {self._yaris_label(r): r["id"] for r in acik_rows}
        self.cb_kayit_yaris["values"] = list(self._yaris_map.keys())
        if self.v_kayit_yaris.get() not in self._yaris_map:
            self.v_kayit_yaris.set("")
        if acik_rows and not self.v_kayit_yaris.get():
            self.v_kayit_yaris.set(self._yaris_label(acik_rows[0]))
        self._kayitlari_listele()

    def _sporculari_yukle(self):
        sec = self.v_kayit_yaris.get().strip()
        yaris_id = self._yaris_map.get(sec) if sec else None
        kayitli = db.yarisa_kayitli_sporcu_idleri(yaris_id) if yaris_id else set()
        rows = db.aktif_lisansli_sporcular()
        self._sporcu_map = {}
        for r in rows:
            if r["sporcu_id"] in kayitli:
                continue
            hib_sporcusu = bool(r["hib_sporcusu"])
            yas_kat = db.hib_kategori_hesapla(r["dogum_tarihi"]) if hib_sporcusu else _yas_kategorisi_hesapla(r["dogum_tarihi"])
            ch = _cinsiyet_harf(r["cinsiyet"])
            gorunen_kategori = f"HİB | {yas_kat}" if hib_sporcusu else yas_kat
            label = f"{r['ad_soyad']} | {gorunen_kategori}({ch}) | {r['lisans_no']} | {r['kulup_adi']}"
            self._sporcu_map[label] = (r["sporcu_id"], r["lisans_id"], r["dogum_tarihi"], r["cinsiyet"], hib_sporcusu)
        self._sporcu_ara_filtrele()
        self._kategori_goruntule()

    def _sporcu_ara_filtrele(self, *_):
        q = self.v_sporcu_ara.get().strip().lower()
        tumu = list(self._sporcu_map.keys())
        if not q:
            filtreli = tumu
        else:
            filtreli = [et for et in tumu if q in et.lower()]
        self.cb_kayit_sporcu["values"] = filtreli
        if filtreli:
            self.v_kayit_sporcu.set(filtreli[0])
        else:
            self.v_kayit_sporcu.set("")
        self._kategori_goruntule()

    def _kategori_goruntule(self, _=None):
        """Seçili sporcunun yaş kategorisini hesapla ve göster."""
        label = self.v_kayit_sporcu.get().strip()
        info = self._sporcu_map.get(label)
        if not info or len(info) < 3:
            self.v_kat_otomatik.set("—")
            return
        kat = db.hib_kategori_hesapla(info[2]) if info[4] else _yas_kategorisi_hesapla(info[2])
        ch = _cinsiyet_harf(info[3])
        self.v_kat_otomatik.set(f"{kat} ({ch})" if kat not in ("—", "KD") else "—")

    def _kayitlari_listele(self):
        sec = self.v_kayit_yaris.get().strip()
        yaris_id = self._yaris_map.get(sec) if sec else None
        rows = db.yaris_kayitlari_listele(yaris_id)

        tree = self.tree_kayit
        tree.delete(*tree.get_children())

        for i, r in enumerate(rows):
            tag = "even" if i % 2 == 0 else "odd"
            ch = _cinsiyet_harf(r["cinsiyet"])
            kat = (db.hib_kategori_hesapla(r["dogum_tarihi"])
                   if r["hib_sporcusu"] else r["kategori"])
            kategori = kat if kat not in ("—", "") else "—"
            tree.insert("", "end", iid=str(r["id"]),
                        values=(r["id"], r["yaris_adi"], r["yaris_tarihi"],
                                r["sporcu"], ch, r["lisans_no"],
                                kategori, r["durum"], r["kayit_tarihi"]),
                        tags=(tag,))

    def _kayit_menu_goster(self, event):
        try:
            self._kayit_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._kayit_menu.grab_release()

    def _on_kayit_cift_tikla(self, event=None):
        sel = self.tree_kayit.selection()
        if not sel:
            return
        self._kayit_guncelle_dialog()

    def _kayit_guncelle_dialog(self):
        """Seçili kaydı güncellemek için küçük bir dialog açar."""
        sel = self.tree_kayit.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Güncellemek için bir kayıt seçin.")
            return
        kayit_id = int(sel[0])

        # Veriyi al
        rows = db.yaris_kayitlari_listele()
        secili = None
        for r in rows:
            if r["id"] == kayit_id:
                secili = r
                break
        if not secili:
            return

        dlg = tk.Toplevel(self)
        dlg.title("Yarış Kaydı Güncelle")
        dlg.configure(bg=BG)
        dlg.geometry("420x320")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()

        frm = ttk.LabelFrame(dlg, text="Kayıt Bilgileri", padding=12)
        frm.pack(fill="both", expand=True, padx=10, pady=10)

        ttk.Label(frm, text="Sporcu:").grid(row=0, column=0, sticky="e", padx=(0, 8), pady=4)
        ttk.Label(frm, text=secili["sporcu"], font=FONT_B).grid(row=0, column=1, sticky="w", pady=4)

        ttk.Label(frm, text="Lisans No:").grid(row=1, column=0, sticky="e", padx=(0, 8), pady=4)
        ttk.Label(frm, text=secili["lisans_no"], font=FONT_B).grid(row=1, column=1, sticky="w", pady=4)

        ttk.Label(frm, text="Yarış:").grid(row=2, column=0, sticky="e", padx=(0, 8), pady=4)
        ttk.Label(frm, text=secili["yaris_adi"], font=FONT_B).grid(row=2, column=1, sticky="w", pady=4)

        ttk.Label(frm, text="Kategori:").grid(row=3, column=0, sticky="e", padx=(0, 8), pady=4)
        v_kat = tk.StringVar(value=secili["kategori"] if secili["kategori"] not in ("—", "") else "")
        cb_kat = ttk.Combobox(frm, textvariable=v_kat,
                              values=["U13", "U15", "U17", "U19", "Elite",
                                      "Master A", "Master B", "Master C", "Diğer"],
                              width=16, state="readonly")
        cb_kat.grid(row=3, column=1, sticky="w", pady=4)

        ttk.Label(frm, text="Durum:").grid(row=4, column=0, sticky="e", padx=(0, 8), pady=4)
        v_dur = tk.StringVar(value=secili["durum"])
        cb_dur = ttk.Combobox(frm, textvariable=v_dur,
                              values=["Onaylandı", "Beklemede", "İptal"],
                              width=14, state="readonly")
        cb_dur.grid(row=4, column=1, sticky="w", pady=4)

        def kaydet():
            kat = v_kat.get().strip() or None
            dur = v_dur.get().strip() or None
            try:
                db.yaris_kayit_guncelle(kayit_id, kategori=kat, durum=dur)
                self._kayitlari_listele()
                dlg.destroy()
                messagebox.showinfo("Başarılı", "Yarış kaydı güncellendi.")
            except Exception as exc:
                messagebox.showerror("Hata", str(exc))

        btn_frm = ttk.Frame(frm)
        btn_frm.grid(row=5, column=0, columnspan=2, pady=(12, 4))
        ttk.Button(btn_frm, text="💾 Kaydet", style="Add.TButton",
                   command=kaydet).pack(side="left", padx=4)
        ttk.Button(btn_frm, text="İptal", style="Neu.TButton",
                   command=dlg.destroy).pack(side="left", padx=4)

    def _kayitlari_csv_export(self):
        """Kayıt listesini CSV olarak dışa aktarır."""
        import csv
        from tkinter import filedialog

        sec = self.v_kayit_yaris.get().strip()
        yaris_id = self._yaris_map.get(sec) if sec else None
        rows = db.yaris_kayitlari_listele(yaris_id)

        yol = filedialog.asksaveasfilename(
            title="Excel/CSV olarak kaydet",
            defaultextension=".csv",
            filetypes=[("CSV Dosyası", "*.csv"), ("Tüm Dosyalar", "*.*")],
            initialfile="yaris_kayitlari.csv",
        )
        if not yol:
            return

        try:
            with open(yol, "w", newline="", encoding="utf-8-sig") as f:
                yazici = csv.writer(f)
                yazici.writerow(["ID", "Yarış", "Yarış Tarihi", "Sporcu",
                                 "Takım", "Cinsiyet", "Lisans No", "Kategori", "Durum", "Kayıt Tarihi"])
                for r in rows:
                    ch = _cinsiyet_harf(r["cinsiyet"])
                    kulup = r["kulup_adi"]
                    yazici.writerow([
                        r["id"], r["yaris_adi"], r["yaris_tarihi"],
                        r["sporcu"], kulup, ch, r["lisans_no"],
                        r["kategori"], r["durum"], r["kayit_tarihi"],
                    ])
            messagebox.showinfo("Başarılı", f"Dışa aktarıldı:\n{yol}")
        except Exception as exc:
            messagebox.showerror("Hata", f"Dışa aktarılamadı: {exc}")

    def _on_kayit_yaris_sec(self, _=None):
        self._kayitlari_listele()
        self._sporculari_yukle()

    def _on_yaris_sec(self, _=None):
        sel = self.tree_yaris.selection()
        if not sel:
            return
        self._secili_yaris_id = int(sel[0])
        row = db.yaris_getir(self._secili_yaris_id)
        if not row:
            return
        self.v_yaris_ad.set(row["ad"] or "")
        self.v_yaris_tarih.set(row["tarih"] or "")
        self.v_yaris_yer.set(row["yer"] or "")
        self.v_yaris_disiplin.set(row["disiplin"] or "Diğer")
        self.v_yaris_sezon.set(row["sezon"] or "2026")
        self.v_yaris_durum.set(row["durum"] or "Planlandı")
        label = self._yaris_label(row)
        if row["durum"] == "Kayıt Açık" and label in self._yaris_map:
            self.v_kayit_yaris.set(label)
            self._kayitlari_listele()
            self._sporculari_yukle()

    def _yaris_ekle(self):
        ad = self.v_yaris_ad.get().strip()
        sezon = self.v_yaris_sezon.get().strip()
        if not ad or not sezon:
            messagebox.showwarning("Uyarı", "Yarış adı ve sezon zorunludur.")
            return
        try:
            db.yaris_ekle(
                ad,
                sezon=sezon,
                tarih=self.v_yaris_tarih.get() or None,
                yer=self.v_yaris_yer.get() or None,
                disiplin=self.v_yaris_disiplin.get() or "Diğer",
                durum=self.v_yaris_durum.get() or "Planlandı",
            )
            self.v_yaris_ad.set("")
            self.v_yaris_tarih.set("")
            self.v_yaris_yer.set("")
            self.v_yaris_disiplin.set("Yol")
            self.v_yaris_durum.set("Kayıt Açık")
            self._yarislari_yukle()
            messagebox.showinfo("Başarılı", "Yarış kaydı eklendi.")
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    def _yaris_guncelle(self):
        if not self._secili_yaris_id:
            messagebox.showwarning("Uyarı", "Listeden bir yarış seçin.")
            return
        ad = self.v_yaris_ad.get().strip()
        sezon = self.v_yaris_sezon.get().strip()
        if not ad or not sezon:
            messagebox.showwarning("Uyarı", "Yarış adı ve sezon zorunludur.")
            return
        try:
            db.yaris_guncelle(
                self._secili_yaris_id,
                ad=ad,
                tarih=self.v_yaris_tarih.get() or None,
                yer=self.v_yaris_yer.get() or None,
                disiplin=self.v_yaris_disiplin.get() or "Diğer",
                sezon=sezon,
                durum=self.v_yaris_durum.get() or "Planlandı",
            )
            self._yarislari_yukle()
            messagebox.showinfo("Başarılı", "Yarış güncellendi.")
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    def _yaris_sil(self):
        if not self._secili_yaris_id:
            messagebox.showwarning("Uyarı", "Listeden bir yarış seçin.")
            return
        if not messagebox.askyesno("Onay", "Seçili yarış silinsin mi?"):
            return
        try:
            db.yaris_sil(self._secili_yaris_id)
            self._secili_yaris_id = None
            self.v_yaris_ad.set("")
            self.v_yaris_tarih.set("")
            self.v_yaris_yer.set("")
            self.v_yaris_disiplin.set("Yol")
            self.v_yaris_sezon.set("2026")
            self.v_yaris_durum.set("Kayıt Açık")
            self._yarislari_yukle()
            messagebox.showinfo("Başarılı", "Yarış silindi.")
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    def _kayit_ekle(self):
        yaris = self.v_kayit_yaris.get().strip()
        sporcu = self.v_kayit_sporcu.get().strip()
        yaris_id = self._yaris_map.get(yaris)
        sporcu_info = self._sporcu_map.get(sporcu)
        if not yaris_id or not sporcu_info:
            messagebox.showwarning(
                "Uyarı",
                "Kayıt için sadece 'Kayıt Açık' yarış seçilebilir ve sporcu seçimi zorunludur.",
            )
            return
        sporcu_id, lisans_id, dogum_tarihi, _, hib_sporcusu = sporcu_info
        kategori = "HİB" if hib_sporcusu else _yas_kategorisi_hesapla(dogum_tarihi)
        if kategori in ("—", "KD"):
            kategori = None
        try:
            db.yaris_kayit_ekle(
                yaris_id=yaris_id,
                sporcu_id=sporcu_id,
                lisans_id=lisans_id,
                kategori=kategori,
            )
            # Optimize: eklenen sporcuyu yerel haritadan kaldır, yeniden sorgulama
            self._sporcu_map.pop(sporcu, None)
            self._sporcu_ara_filtrele()
            self._kayitlari_listele()
            messagebox.showinfo("Başarılı", "Sporcu yarışa kaydedildi.")
        except Exception as exc:
            if "UNIQUE constraint failed" in str(exc):
                messagebox.showwarning(
                    "Uyarı",
                    "Bu sporcu seçili yarışa zaten kayıtlı.",
                )
                return
            messagebox.showerror("Hata", str(exc))

    def _kayit_sil(self):
        sel = self.tree_kayit.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Silmek için bir kayıt seçin.")
            return
        kayit_id = int(sel[0])
        if not messagebox.askyesno("Onay", "Seçili yarış kaydı silinsin mi?"):
            return
        db.yaris_kayit_sil(kayit_id)
        self._kayitlari_listele()
        self._sporculari_yukle()
        messagebox.showinfo("Başarılı", "Yarış kaydı silindi.")

    def _on_yaris_cift_tikla(self, event=None):
        sel = self.tree_yaris.selection()
        if not sel:
            return
        self._kayit_penceresi_ac(yaris_id=int(sel[0]))

    def _kayit_penceresi_ac(self, yaris_id: int | None = None):  # noqa: C901
        import datetime

        # ── Kategori sırası ve bir-alt haritası ──────────────────────────
        KATEGORI_SIRA = [
            "U13", "U15", "U17", "U19",
            "Elite", "Master A", "Master B", "Master C",
        ]

        def hesapla_yas_kat(dogum_tarihi):
            if not dogum_tarihi:
                return "—"
            try:
                yas = datetime.date.today().year - int(dogum_tarihi[:4])
                if 11 <= yas <= 12:  return "U13"
                if 13 <= yas <= 14:  return "U15"
                if 15 <= yas <= 16:  return "U17"
                if 17 <= yas <= 18:  return "U19"
                if 19 <= yas <= 34:  return "Elite"
                if 35 <= yas <= 39:  return "Master A"
                if 40 <= yas <= 44:  return "Master B"
                if yas >= 45:        return "Master C"
                return "Kategori Dışı"
            except Exception:
                return "—"

        # ── Pencere ───────────────────────────────────────────────────────
        win = tk.Toplevel(self)
        win.title("Yarışa Sporcu Kayıt")
        win.configure(bg=BG)
        _apply_window_geometry(
            win,
            width_ratio=0.9,
            height_ratio=0.88,
            min_width=1200,
            min_height=780,
        )

        tum_yarislar = db.yarislar_listele()
        popup_yaris_map  = {self._yaris_label(r): r["id"]     for r in tum_yarislar}
        yaris_sezon_map  = {r["id"]: (r["sezon"] or "")       for r in tum_yarislar}
        yaris_disiplin_map = {r["id"]: (r["disiplin"] or "") for r in tum_yarislar}

        if yaris_id is not None:
            row0 = db.yaris_getir(yaris_id)
            baslangic_label = self._yaris_label(row0) if row0 else self.v_kayit_yaris.get()
        else:
            baslangic_label = self.v_kayit_yaris.get()

        # ── Form (grid) ───────────────────────────────────────────────────
        top = ttk.LabelFrame(win, text="Kayıt Bilgileri", padding=8)
        top.pack(fill="x", padx=8, pady=8)

        # Satır 0: Yarış | Kayıt Durumu
        ttk.Label(top, text="Yarış:").grid(
            row=0, column=0, sticky="e", padx=(4, 4), pady=4)
        v_yaris = tk.StringVar(value=baslangic_label)
        cb = ttk.Combobox(top, textvariable=v_yaris,
                          values=list(popup_yaris_map.keys()),
                          state="readonly", width=40)
        cb.grid(row=0, column=1, columnspan=3, sticky="ew", padx=(0, 8), pady=4)
        ttk.Label(top, text="Kayıt Durumu:").grid(
            row=0, column=4, sticky="e", padx=(4, 4), pady=4)
        v_durum = tk.StringVar(value="Onaylandı")
        cb_durum = ttk.Combobox(top, textvariable=v_durum,
                                values=["Onaylandı", "Beklemede", "İptal"],
                                state="readonly", width=12)
        cb_durum.grid(row=0, column=5, sticky="ew", padx=(0, 8), pady=4)

        # Satır 1: Sporcu + Arama
        ttk.Label(top, text="Sporcu:").grid(
            row=1, column=0, sticky="e", padx=(4, 4), pady=4)
        v_sporcu_ara = tk.StringVar()
        e_sporcu_ara = ttk.Entry(top, textvariable=v_sporcu_ara, width=24)
        e_sporcu_ara.grid(row=1, column=1, sticky="ew", padx=(0, 4), pady=4)
        ttk.Label(top, text="Ara →").grid(
            row=1, column=2, sticky="w", padx=(0, 2), pady=4)
        v_sporcu = tk.StringVar()
        cb_sporcu = ttk.Combobox(top, textvariable=v_sporcu,
                                 state="readonly", width=46)
        cb_sporcu.grid(row=1, column=3, columnspan=3,
                       sticky="ew", padx=(0, 8), pady=4)

        # Satır 1b: Sporcu seçim kısayol butonları
        def _sporcu_ara_filtrele(*_):
            """Arama çubuğundaki metne göre sporcu combobox'ını filtreler."""
            q = v_sporcu_ara.get().strip().lower()
            tumu = _sporcu_ara_tum_liste
            if not q:
                cb_sporcu["values"] = tumu
                if tumu:
                    v_sporcu.set(tumu[0])
                return
            filtreli = [et for et in tumu if q in et.lower()]
            cb_sporcu["values"] = filtreli
            if filtreli:
                v_sporcu.set(filtreli[0])
            else:
                v_sporcu.set("")

        _sporcu_ara_tum_liste: list = []

        def refresh_sporcular(secili_yaris_id):
            nonlocal sporcu_map, _sporcu_ara_tum_liste
            kayitli = (db.yarisa_kayitli_sporcu_idleri(secili_yaris_id)
                       if secili_yaris_id else set())
            rows = db.aktif_lisansli_sporcular()
            sporcu_map = {}
            for r in rows:
                if r["sporcu_id"] in kayitli:
                    continue
                hib_sporcusu = bool(r["hib_sporcusu"])
                yas_kat = db.hib_kategori_hesapla(r["dogum_tarihi"]) if hib_sporcusu else _yas_kategorisi_hesapla(r["dogum_tarihi"])
                ch = _cinsiyet_harf(r["cinsiyet"])
                gorunen_kategori = f"HİB | {yas_kat}" if hib_sporcusu else yas_kat
                label = f"{r['ad_soyad']} | {gorunen_kategori}({ch}) | {r['lisans_no']} | {r['kulup_adi']}"
                sporcu_map[label] = (r["sporcu_id"], r["lisans_id"], r["dogum_tarihi"], r["cinsiyet"], hib_sporcusu)
            _sporcu_ara_tum_liste = list(sporcu_map.keys())
            _sporcu_ara_filtrele()
            on_sporcu_sec()

        v_sporcu_ara.trace_add("write", _sporcu_ara_filtrele)

        # Satır 2: Yaş Kategorisi | Yarış Kategorisi
        ttk.Label(top, text="Yaş Kategorisi:").grid(
            row=2, column=0, sticky="e", padx=(4, 4), pady=4)
        lbl_yas_kat = ttk.Label(top, text="—", font=FONT_B)
        lbl_yas_kat.grid(row=2, column=1, sticky="w", padx=(0, 16), pady=4)
        ttk.Label(top, text="Yarış Kategorisi:").grid(
            row=2, column=2, sticky="e", padx=(4, 4), pady=4)
        v_yaris_kat = tk.StringVar(value="—")
        lbl_yaris_kat = ttk.Label(top, textvariable=v_yaris_kat, font=FONT_B)
        lbl_yaris_kat.grid(row=2, column=3, sticky="w", padx=(0, 8), pady=4)

        # Satır 3: İşlem butonları
        btn_frame = ttk.Frame(top)
        btn_frame.grid(row=3, column=0, columnspan=6, pady=(8, 2))
        top.columnconfigure(1, weight=1)

        # ── Kayıt listesi ─────────────────────────────────────────────────
        cols = [
            ("id",           "ID",           42),
            ("sporcu",       "Sporcu",       230),
            ("cinsiyet",     "Cinsiyet",      55),
            ("lisans_no",    "Lisans No",     95),
            ("kategori",     "Kategori",     110),
            ("durum",        "Durum",         90),
            ("kayit_tarihi", "Kayıt Tarihi", 100),
        ]
        tf, tree = _make_tree(win, cols)
        tf.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # ── İç durum ──────────────────────────────────────────────────────
        sporcu_map: dict = {}   # label → (sporcu_id, lisans_id, dogum_tarihi, cinsiyet)
        _st = {"yas_kat": "—", "override_kat": None}

        # ── Yardımcı fonksiyonlar ─────────────────────────────────────────
        def get_secili_sezon():
            yid = popup_yaris_map.get(v_yaris.get().strip())
            return yaris_sezon_map.get(yid, "") if yid else ""

        def get_secili_disiplin():
            yid = popup_yaris_map.get(v_yaris.get().strip())
            return yaris_disiplin_map.get(yid, "Yol") if yid else "Yol"

        def get_sporcu_yaris_kategorisi(sporcu_id):
            sezon = get_secili_sezon()
            kategori = (
                db.sporcu_sezon_kayitli_kategori(
                    sporcu_id, sezon, get_secili_disiplin()
                )
                if sezon else None
            )
            return kategori or _st["yas_kat"]

        def update_kat_display():
            yas = _st["yas_kat"]
            ov  = _st["override_kat"]
            if ov and ov != yas:
                v_yaris_kat.set(ov)
                lbl_yaris_kat.configure(foreground=ACCENT)
            else:
                v_yaris_kat.set(yas)
                lbl_yaris_kat.configure(foreground="black")

        def on_sporcu_sec(_=None):
            info = sporcu_map.get(v_sporcu.get().strip())
            if not info:
                _st["yas_kat"] = "—"
                _st["override_kat"] = None
                lbl_yas_kat.config(text="—")
                update_kat_display()
                return
            sporcu_id, _lid, dogum_tarihi, cinsiyet, hib_sporcusu = info
            yas_kat = db.hib_kategori_hesapla(dogum_tarihi) if hib_sporcusu else hesapla_yas_kat(dogum_tarihi)
            ch = _cinsiyet_harf(cinsiyet)
            _st["yas_kat"] = yas_kat
            lbl_yas_kat.config(text=f"{yas_kat} ({ch})")
            if hib_sporcusu:
                _st["override_kat"] = None
                update_kat_display()
                return
            # Bu sezonda zaten kayıtlı kategorisi var mı?
            sezon = get_secili_sezon()
            ov = (
                db.sporcu_sezon_kayitli_kategori(
                    sporcu_id, sezon, get_secili_disiplin()
                )
                if sezon else None
            )
            _st["override_kat"] = ov if (ov and ov != yas_kat) else None
            update_kat_display()

        cb_sporcu.bind("<<ComboboxSelected>>", on_sporcu_sec)



        def fill_tree_kayit(rows):
            tree.delete(*tree.get_children())
            for i, row in enumerate(rows):
                tag = "even" if i % 2 == 0 else "odd"
                ch = _cinsiyet_harf(row["cinsiyet"])
                kat = (db.hib_kategori_hesapla(row["dogum_tarihi"])
                       if row["hib_sporcusu"] else row["kategori"])
                kategori_goruntule = kat if kat not in ("—", "") else "—"
                vals = (
                    row["id"], row["sporcu"], ch,
                    row["lisans_no"], kategori_goruntule,
                    row["durum"], row["kayit_tarihi"],
                )
                tree.insert("", "end", iid=str(row["id"]),
                            values=vals, tags=(tag,))

        def refresh(_=None):
            sec = v_yaris.get().strip()
            secili_yaris_id = popup_yaris_map.get(sec) if sec else None
            rows = db.yaris_kayitlari_listele(secili_yaris_id)
            fill_tree_kayit(rows)
            refresh_sporcular(secili_yaris_id)

        def kayit_ekle_popup():
            sec = v_yaris.get().strip()
            ekle_yaris_id = popup_yaris_map.get(sec) if sec else None
            secili_label = v_sporcu.get().strip()
            sporcu_info = sporcu_map.get(secili_label)
            if not ekle_yaris_id or not sporcu_info:
                messagebox.showwarning(
                    "Uyarı",
                    "Kayıt için yarış ve sporcu seçimi zorunludur.",
                    parent=win,
                )
                return
            sporcu_id, lisans_id, _, _, hib_sporcusu = sporcu_info
            kat = "HİB" if hib_sporcusu else get_sporcu_yaris_kategorisi(sporcu_id)
            kategori = None if kat in ("—", "Kategori Dışı") else kat
            try:
                db.yaris_kayit_ekle(
                    yaris_id=ekle_yaris_id,
                    sporcu_id=sporcu_id,
                    lisans_id=lisans_id,
                    kategori=kategori,
                )
                # Optimize: yerel haritadan kaldır, yeniden sorgulama
                sporcu_map.pop(secili_label, None)
                if secili_label in _sporcu_ara_tum_liste:
                    _sporcu_ara_tum_liste.remove(secili_label)
                _sporcu_ara_filtrele()
                on_sporcu_sec()
                fill_tree_kayit(
                    db.yaris_kayitlari_listele(ekle_yaris_id))
                messagebox.showinfo("Başarılı", "Sporcu yarışa kaydedildi.", parent=win)
            except Exception as exc:
                if "UNIQUE constraint failed" in str(exc):
                    messagebox.showwarning(
                        "Uyarı", "Bu sporcu bu yarışa zaten kayıtlı.", parent=win)
                    return
                messagebox.showerror("Hata", str(exc), parent=win)

        def kayit_sil_popup():
            sel_items = tree.selection()
            if not sel_items:
                messagebox.showwarning("Uyarı", "Silmek için bir kayıt seçin.", parent=win)
                return
            kayit_id = int(sel_items[0])
            if not messagebox.askyesno("Onay", "Seçili yarış kaydı silinsin mi?", parent=win):
                return
            db.yaris_kayit_sil(kayit_id)
            refresh()
            messagebox.showinfo("Başarılı", "Yarış kaydı silindi.", parent=win)

        def kayit_guncelle_popup():
            sel_items = tree.selection()
            if not sel_items:
                messagebox.showwarning("Uyarı", "Güncellemek için bir kayıt seçin.", parent=win)
                return
            kayit_id = int(sel_items[0])
            kat = _st["override_kat"] or _st["yas_kat"]
            kategori = None if kat in ("—", "Kategori Dışı") else kat
            try:
                db.yaris_kayit_guncelle(
                    kayit_id,
                    kategori=kategori,
                    durum=v_durum.get() or None,
                )
                refresh()
                messagebox.showinfo("Başarılı", "Yarış kaydı güncellendi.", parent=win)
            except Exception as exc:
                messagebox.showerror("Hata", str(exc), parent=win)

        def on_tree_sec(_=None):
            sel_items = tree.selection()
            if not sel_items:
                return
            vals = tree.item(sel_items[0], "values")
            # vals: (id, sporcu, cinsiyet, lisans_no, kategori, durum, kayit_tarihi)
            if len(vals) >= 6:
                v_durum.set(vals[5] or "Onaylandı")

        def kayit_sec_forma_yukle():
            """Seçili kaydın bilgilerini form alanlarına yükler."""
            sel_items = tree.selection()
            if not sel_items:
                return
            kayit_id = int(sel_items[0])
            rows = db.yaris_kayitlari_listele()
            secili = next((r for r in rows if r["id"] == kayit_id), None)
            if not secili:
                return
            # Yarış seçimini güncelle (popup_yaris_map key: "ad (tarih)")
            yaris_tarih = secili["yaris_tarihi"] if secili["yaris_tarihi"] not in ("—", "") else "Tarihsiz"
            yaris_label = f"{secili['yaris_adi']} ({yaris_tarih})"
            if yaris_label in popup_yaris_map:
                v_yaris.set(yaris_label)
            # Durum
            v_durum.set(secili["durum"] or "Onaylandı")
            # Yaş kategorisi (cinsiyet ile birlikte)
            ch = _cinsiyet_harf(secili["cinsiyet"])
            yas_kat = hesapla_yas_kat(secili["dogum_tarihi"])
            lbl_yas_kat.config(text=f"{yas_kat} ({ch})" if yas_kat != "—" else "—")
            _st["yas_kat"] = yas_kat
            # Kategori override varsa göster, yoksa yaş kategorisini kullan
            db_kat = secili["kategori"]
            if db_kat not in ("—", "", None) and db_kat != yas_kat:
                _st["override_kat"] = db_kat
            else:
                _st["override_kat"] = None
            update_kat_display()

        def kayit_cift_tikla(_=None):
            if not tree.selection():
                return
            kayit_sec_forma_yukle()

        tree.bind("<<TreeviewSelect>>", on_tree_sec)
        tree.bind("<Double-1>", kayit_cift_tikla)

        ttk.Button(btn_frame, text="➕ Kaydı Oluştur", style="Add.TButton",
                   command=kayit_ekle_popup).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="✏️ Seçili Kaydı Güncelle", style="Upd.TButton",
                   command=kayit_guncelle_popup).pack(side="left", padx=(0, 4))
        ttk.Button(btn_frame, text="✖ Seçili Kaydı Sil", style="Del.TButton",
                   command=kayit_sil_popup).pack(side="left", padx=(0, 4))

        def _popup_csv_export():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from tkinter import filedialog
            sec = v_yaris.get().strip()
            secili_yaris_id = popup_yaris_map.get(sec) if sec else None
            rows = db.yaris_kayitlari_listele(secili_yaris_id)
            if not rows:
                messagebox.showinfo("Bilgi", "Dışa aktarılacak kayıt bulunamadı.", parent=win)
                return
            yol = filedialog.asksaveasfilename(
                title="Excel olarak kaydet (Gruplu)",
                defaultextension=".xlsx",
                filetypes=[("Excel Dosyası", "*.xlsx")],
                initialfile="yaris_kayitlari_gruplu.xlsx",
                parent=win,
            )
            if not yol:
                return
            from itertools import groupby
            satirlar = []
            for r in rows:
                ch = _cinsiyet_harf(r["cinsiyet"])
                kat = r["kategori"] if r["kategori"] not in ("—", "") else "—"
                kulup = r["kulup_adi"]
                satirlar.append({
                    "id": r["id"], "sporcu": r["sporcu"], "cinsiyet": ch,
                    "lisans_no": r["lisans_no"], "takim": kulup,
                    "kategori": f"{kat} ({ch})" if kat != "—" else "—",
                    "durum": r["durum"], "kayit_tarihi": r["kayit_tarihi"],
                    "_cinsiyet_raw": r["cinsiyet"], "_kategori_raw": r["kategori"],
                })
            def anahtar(s):
                k = s["_kategori_raw"] if s["_kategori_raw"] not in ("—", "") else "Belirtilmemiş"
                c = _cinsiyet_harf(s["_cinsiyet_raw"])
                return (k, c)
            satirlar.sort(key=anahtar)

            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Yarış Kayıtları"

                # Başlık stilleri
                kalin = Font(bold=True, size=11)
                grup_font = Font(bold=True, size=11, color="FFFFFF")
                grup_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
                baslik_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
                baslik_font = Font(bold=True, size=11, color="FFFFFF")

                # Başlık satırı
                basliklar = ["Kategori", "Cinsiyet", "ID", "Sporcu", "Takım", "Lisans No", "Durum", "Kayıt Tarihi"]
                for col, b in enumerate(basliklar, 1):
                    h = ws.cell(row=1, column=col, value=b)
                    h.font = baslik_font
                    h.fill = baslik_fill
                    h.alignment = Alignment(horizontal="center")

                satir_no = 2
                for g, elemanlar in groupby(satirlar, key=anahtar):
                    kat, ch = g
                    el = list(elemanlar)
                    # Grup başlık satırı
                    ws.merge_cells(start_row=satir_no, start_column=1, end_row=satir_no, end_column=8)
                    h = ws.cell(row=satir_no, column=1,
                                value=f"{kat} ({ch}) — {len(el)} kayıt")
                    h.font = grup_font
                    h.fill = grup_fill
                    satir_no += 1
                    # Sporcu satırları
                    for s in el:
                        ws.cell(row=satir_no, column=1, value="")
                        ws.cell(row=satir_no, column=2, value=ch)
                        ws.cell(row=satir_no, column=3, value=s["id"])
                        ws.cell(row=satir_no, column=4, value=s["sporcu"])
                        ws.cell(row=satir_no, column=5, value=s["takim"])
                        ws.cell(row=satir_no, column=6, value=s["lisans_no"])
                        ws.cell(row=satir_no, column=7, value=s["durum"])
                        ws.cell(row=satir_no, column=8, value=s["kayit_tarihi"])
                        satir_no += 1

                # Sütun genişlikleri
                ws.column_dimensions['A'].width = 22
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 8
                ws.column_dimensions['D'].width = 28
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 14
                ws.column_dimensions['G'].width = 14
                ws.column_dimensions['H'].width = 16

                wb.save(yol)
                messagebox.showinfo("Başarılı", f"Excel dosyası kaydedildi:\n{yol}", parent=win)
            except Exception as exc:
                messagebox.showerror("Hata", f"Dışa aktarılamadı: {exc}", parent=win)

        ttk.Button(btn_frame, text="📊 Excel (Gruplu)", style="Neu.TButton",
                   command=_popup_csv_export).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="🔄 Yenile", style="Neu.TButton",
                   command=refresh).pack(side="left", padx=4)

        cb.bind("<<ComboboxSelected>>", refresh)
        refresh()


# ---------------------------------------------------------------------------
# EVRAK KONTROL sekmesi
# ---------------------------------------------------------------------------

class EvrakKontrolSekme(ttk.Frame):
    BASVURU_TURLERI = {
        "Kulüp Üyelik / Aktif Üyelik": {
            "referans_turu": "kulup",
            "kaynak": "KT_Bisiklet_Federasyonu_Üyelik_Talimatı ve kulüp başvuru formu",
            "formlar": [
                ("EK-K1", "Kulüp yeniden kayıt / aktif üyelik başvuru formu",
                 "KULÜP YENİDEN KAYIT / AKTİF ÜYELİK BAŞVURU FORMU"),
                ("EK-K2", "Kulüp yetkili ve iletişim bilgileri formu",
                 "KULÜP YETKİLİ VE İLETİŞİM BİLGİLERİ FORMU"),
                ("EK-K3", "Kulüp tescil fişi", "KULÜP TESCİL FİŞİ"),
            ],
            "belgeler": [
                ("uyelik_formu", "Kulüp yeniden kayıt / aktif üyelik başvuru formu", 1),
                ("yetkili_bilgileri", "Kulüp yetkili ve iletişim bilgileri", 1),
                ("aidat_teyidi", "Aidat ödeme / üyelik durumu teyidi", 1),
            ],
        },
        "Sporcu Lisans Başvurusu": {
            "referans_turu": "sporcu",
            "kaynak": "Sporcu_Lisans_Talimatı Madde 7, 7A, 8 ve EK-1",
            "formlar": [
                ("EK-1", "Sporcu lisans başvuru formu", "SPORCU LİSANS BAŞVURU FORMU"),
                ("EK-7", "Veli muvafakatname örneği", "EK-7 VELİ MUVAFAKATNAME FORMU"),
                ("EK-VI1", "Veli izni şablonu", "VELİ İZİN FORMU"),
            ],
            "belgeler": [
                ("spor_dairesi_kaydi", "Spor Dairesi BYS kaydı", 1),
                ("saglik_raporu", "Sağlık raporu", 1),
                ("veli_muvafakati", "Veli muvafakatnamesi", 0),
                ("baska_federasyon_beyani", "Başka federasyon lisans beyanı", 0),
            ],
        },
        "Vize / Lisans Yenileme": {
            "referans_turu": "sporcu",
            "kaynak": "Sporcu_Lisans_Talimatı EK-4",
            "formlar": [
                ("EK-4", "Vize / lisans yenileme formu", "VİZE / LİSANS YENİLEME BAŞVURU FORMU"),
            ],
            "belgeler": [
                ("saglik_raporu", "Sağlık raporu", 1),
                ("eski_lisans_teslim", "Eski lisans teslimi", 1),
                ("kulup_yetkilisi", "Kulüp yetkilisi onayı / bilgisi", 0),
            ],
        },
        "Transfer Başvurusu": {
            "referans_turu": "sporcu",
            "kaynak": "Sporcu_Lisans_Talimatı Madde 9",
            "formlar": [
                ("TR-1", "Transfer talep formu", "SPORCU TRANSFER TALEP FORMU"),
                ("TR-2", "İlişiksizlik belgesi örneği", "İLİŞİKSİZLİK BELGESİ"),
            ],
            "belgeler": [
                ("ilizsizlik_belgesi", "İlişiksizlik belgesi", 1),
            ],
        },
        "Yurt Dışı Yarış İzni": {
            "referans_turu": "sporcu",
            "kaynak": "Sporcu_Lisans_Talimatı Madde 10-11 ve EK-3",
            "formlar": [
                ("EK-3", "Yurt dışı yarış izin başvuru formu", "YURT DIŞI YARIŞ İZİN BAŞVURU FORMU"),
            ],
            "belgeler": [
                ("kulup_yazisi", "Kulüp yazısı / organizasyon onayı", 1),
            ],
        },
        "Yabancı Uyruklu / Misafir Sporcu": {
            "referans_turu": "sporcu",
            "kaynak": "Yabanci_Uyruklu_Misafir_Sporcu_Talimati",
            "formlar": [
                ("YU-1", "Yabancı uyruklu / misafir sporcu başvuru formu",
                 "YABANCI UYRUKLU / MİSAFİR SPORCU BAŞVURU FORMU"),
            ],
            "belgeler": [
                ("pasaport_kimlik", "Pasaport / kimlik belgesi", 1),
                ("yabanci_lisans", "Yabancı federasyon lisans bilgisi", 1),
                ("gecerlilik_teyidi", "Lisans geçerlilik teyidi", 1),
            ],
        },
        "Yabancı Federasyon Lisanslı KKTC Vatandaşı": {
            "referans_turu": "sporcu",
            "kaynak": "KTBF_Yabanci_Federasyon_Lisansli_KKTC_Vatandasi_Sporcular_Talimati",
            "formlar": [
                ("YF-1", "Yabancı federasyon lisanslı KKTC vatandaşı sporcu formu",
                 "YABANCI FEDERASYON LİSANSLI KKTC VATANDAŞI SPORCU BAŞVURU FORMU"),
                ("YF-2", "Kulüp muvafakatname örneği", "KULÜP MUVAFAKATNAME ÖRNEĞİ"),
            ],
            "belgeler": [
                ("yabanci_lisans", "Yabancı federasyon lisans bilgisi", 1),
                ("gecerlilik_tarihi", "Lisans geçerlilik tarihi teyidi", 1),
                ("kulup_muvafakati", "Kulüp muvafakati", 0),
            ],
        },
    }

    def __init__(self, parent):
        super().__init__(parent, style="TFrame")
        self._sporcu_map: dict[str, int] = {}
        self._kulup_map: dict[str, int] = {}
        self._evrak_satirlari: dict[str, dict] = {}
        self._ozet_map: dict[str, dict] = {}
        self._aktif_form_map: dict[str, dict] = {}
        self._son_form_dosyasi = ""
        self._build()
        self._referanslari_yukle()
        self._basvuru_turlerini_yukle()
        self._ozet_listele()

    def _build(self):
        ttk.Label(self, text="EVRAK KONTROL MODÜLÜ",
                  style="Header.TLabel").pack(fill="x")

        ust = ttk.LabelFrame(self, text="Başvuru Seçimi", padding=8)
        ust.pack(fill="x", padx=8, pady=(8, 4))
        ust.columnconfigure(1, weight=1)
        ust.columnconfigure(3, weight=1)

        ttk.Label(ust, text="Başvuru Türü").grid(
            row=0, column=0, sticky="e", padx=(8, 4), pady=4)
        self.v_basvuru = tk.StringVar()
        self.cb_basvuru = ttk.Combobox(
            ust,
            textvariable=self.v_basvuru,
            state="readonly",
            width=34,
        )
        self.cb_basvuru.grid(row=0, column=1, sticky="ew", padx=(0, 8), pady=4)
        self.cb_basvuru.bind("<<ComboboxSelected>>", lambda _=None: self._basvuru_degisti())

        ttk.Label(ust, text="Sezon").grid(
            row=0, column=2, sticky="e", padx=(8, 4), pady=4)
        self.v_sezon = tk.StringVar(value="2026")
        ttk.Entry(ust, textvariable=self.v_sezon, width=12).grid(
            row=0, column=3, sticky="w", padx=(0, 8), pady=4)

        ttk.Label(ust, text="Sporcu").grid(
            row=1, column=0, sticky="e", padx=(8, 4), pady=4)
        self.v_sporcu = tk.StringVar()
        self.cb_sporcu = ttk.Combobox(ust, textvariable=self.v_sporcu,
                                      state="readonly", width=34)
        self.cb_sporcu.grid(row=1, column=1, sticky="ew", padx=(0, 8), pady=4)

        ttk.Label(ust, text="Kulüp").grid(
            row=1, column=2, sticky="e", padx=(8, 4), pady=4)
        self.v_kulup = tk.StringVar()
        self.cb_kulup = ttk.Combobox(ust, textvariable=self.v_kulup,
                                     state="readonly", width=28)
        self.cb_kulup.grid(row=1, column=3, sticky="ew", padx=(0, 8), pady=4)

        ttk.Label(ust, text="Talimat Kaynağı").grid(
            row=2, column=0, sticky="ne", padx=(8, 4), pady=4)
        self.lbl_kaynak = ttk.Label(ust, text="—", justify="left")
        self.lbl_kaynak.grid(row=2, column=1, columnspan=3,
                             sticky="w", padx=(0, 8), pady=4)

        ttk.Label(ust, text="Talimat Eki Formu").grid(
            row=3, column=0, sticky="e", padx=(8, 4), pady=4)
        self.v_form_ornek = tk.StringVar()
        self.cb_form_ornek = ttk.Combobox(
            ust,
            textvariable=self.v_form_ornek,
            state="readonly",
            width=34,
            values=[],
        )
        self.cb_form_ornek.grid(row=3, column=1, sticky="ew", padx=(0, 8), pady=4)

        frm_actions = ttk.Frame(ust)
        frm_actions.grid(row=3, column=2, columnspan=2, sticky="w", pady=4)
        ttk.Button(frm_actions, text="🧾 Form Üret", style="Neu.TButton",
                   command=self._form_onizleme_ac).pack(side="left", padx=4)
        ttk.Button(frm_actions, text="🖨 Form Yazdır", style="Neu.TButton",
                   command=self._formu_yazdir).pack(side="left", padx=4)

        aksiyon = ttk.Frame(ust)
        aksiyon.grid(row=4, column=0, columnspan=4, sticky="w", pady=(8, 0))
        ttk.Button(aksiyon, text="💾 Kontrolü Kaydet", style="Add.TButton",
                   command=self._kaydet).pack(side="left", padx=4)
        ttk.Button(aksiyon, text="🔄 Kayıtları Yükle", style="Neu.TButton",
                   command=self._mevcut_kaydi_yukle).pack(side="left", padx=4)
        ttk.Button(aksiyon, text="🗂 Listeleri Yenile", style="Neu.TButton",
                   command=self._referanslari_yukle).pack(side="left", padx=4)
        ttk.Button(aksiyon, text="✖ Temizle", style="Neu.TButton",
                   command=self._temizle).pack(side="left", padx=4)

        self.frm_evrak = ttk.LabelFrame(self, text="İstenen Evraklar", padding=8)
        self.frm_evrak.pack(fill="x", padx=8, pady=4)
        self.frm_evrak.columnconfigure(2, weight=1)

        alt = ttk.LabelFrame(self, text="Kayıtlı Evrak Kontrolleri", padding=4)
        alt.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        cols = [
            ("basvuru", "Başvuru", 240),
            ("referans", "Referans", 200),
            ("sezon", "Sezon", 70),
            ("tamam", "Tamam", 80),
            ("guncelleme", "Güncelleme", 95),
        ]
        tf, self.tree_ozet = _make_tree(alt, cols)
        tf.pack(fill="both", expand=True)
        self.tree_ozet.bind("<<TreeviewSelect>>", self._on_ozet_sec)

    def _basvuru_turlerini_yukle(self):
        basvurular = list(self.BASVURU_TURLERI.keys())
        self.cb_basvuru["values"] = basvurular
        if not self.v_basvuru.get() and basvurular:
            self.v_basvuru.set(basvurular[0])
        self._basvuru_degisti()

    def _referanslari_yukle(self):
        sporcular = db.sporcular_dropdown()
        self._sporcu_map = {row["ad_soyad"]: row["id"] for row in sporcular}
        self.cb_sporcu["values"] = list(self._sporcu_map.keys())
        if self.v_sporcu.get() not in self._sporcu_map:
            self.v_sporcu.set("")
        if self._sporcu_map and not self.v_sporcu.get():
            self.v_sporcu.set(next(iter(self._sporcu_map)))

        kulupler = db.kulupler_listele()
        self._kulup_map = {row["ad"]: row["id"] for row in kulupler}
        self.cb_kulup["values"] = list(self._kulup_map.keys())
        if self.v_kulup.get() not in self._kulup_map:
            self.v_kulup.set("")
        if self._kulup_map and not self.v_kulup.get():
            self.v_kulup.set(next(iter(self._kulup_map)))

    def _aktif_kural(self):
        return self.BASVURU_TURLERI[self.v_basvuru.get()]

    def _secili_referans(self):
        referans_turu = self._aktif_kural()["referans_turu"]
        if referans_turu == "sporcu":
            return referans_turu, self._sporcu_map.get(self.v_sporcu.get(), 0)
        if referans_turu == "kulup":
            return referans_turu, self._kulup_map.get(self.v_kulup.get(), 0)
        return referans_turu, 0

    def _basvuru_degisti(self):
        kural = self._aktif_kural()
        referans_turu = kural["referans_turu"]
        self.lbl_kaynak.config(text=kural["kaynak"])
        self.cb_sporcu.configure(state="readonly" if referans_turu == "sporcu" else "disabled")
        self.cb_kulup.configure(state="readonly" if referans_turu == "kulup" else "disabled")
        self._basvuru_formlarini_yukle()
        self._evrak_satirlarini_olustur()
        self._mevcut_kaydi_yukle()

    def _basvuru_formlarini_yukle(self):
        self._aktif_form_map.clear()
        gorunenler = []
        for form_kodu, form_adi, baslik in self._aktif_kural().get("formlar", []):
            gorunen = f"{form_kodu} - {form_adi}"
            gorunenler.append(gorunen)
            self._aktif_form_map[gorunen] = {
                "kod": form_kodu,
                "ad": form_adi,
                "baslik": baslik,
            }
        self.cb_form_ornek["values"] = gorunenler
        if gorunenler:
            self.v_form_ornek.set(gorunenler[0])
            self.cb_form_ornek.configure(state="readonly")
        else:
            self.v_form_ornek.set("")
            self.cb_form_ornek.configure(state="disabled")

    def _evrak_satirlarini_olustur(self):
        for child in self.frm_evrak.winfo_children():
            child.destroy()
        self._evrak_satirlari.clear()

        ttk.Label(self.frm_evrak, text="Belge", font=FONT_B).grid(
            row=0, column=0, sticky="w", padx=(4, 8), pady=(0, 6))
        ttk.Label(self.frm_evrak, text="Durum", font=FONT_B).grid(
            row=0, column=1, sticky="w", padx=(4, 8), pady=(0, 6))
        ttk.Label(self.frm_evrak, text="Not", font=FONT_B).grid(
            row=0, column=2, sticky="w", padx=(4, 8), pady=(0, 6))

        for index, (kod, ad, zorunlu) in enumerate(self._aktif_kural()["belgeler"], start=1):
            ttk.Label(self.frm_evrak, text=ad).grid(
                row=index, column=0, sticky="w", padx=(4, 8), pady=4)
            teslim = tk.IntVar(value=0)
            chk = ttk.Checkbutton(
                self.frm_evrak,
                text="Zorunlu" if zorunlu else "Opsiyonel",
                variable=teslim,
            )
            chk.grid(row=index, column=1, sticky="w", padx=(4, 8), pady=4)
            not_var = tk.StringVar()
            ttk.Entry(self.frm_evrak, textvariable=not_var, width=48).grid(
                row=index, column=2, sticky="ew", padx=(0, 8), pady=4)
            self._evrak_satirlari[kod] = {
                "ad": ad,
                "zorunlu": zorunlu,
                "teslim": teslim,
                "notlar": not_var,
            }
        self.frm_evrak.columnconfigure(2, weight=1)

    def _mevcut_kaydi_yukle(self):
        referans_turu, referans_id = self._secili_referans()
        if referans_turu != "serbest" and not referans_id:
            self._satirlari_sifirla()
            return
        kayitlar = db.evrak_kontrol_getir(
            self.v_basvuru.get(),
            referans_turu,
            referans_id=referans_id,
            sezon=self.v_sezon.get().strip(),
        )
        self._satirlari_sifirla()
        for kayit in kayitlar:
            satir = self._evrak_satirlari.get(kayit["belge_kodu"])
            if not satir:
                continue
            satir["teslim"].set(kayit["teslim"])
            satir["notlar"].set(kayit["notlar"] or "")

    def _satirlari_sifirla(self):
        for satir in self._evrak_satirlari.values():
            satir["teslim"].set(0)
            satir["notlar"].set("")

    def _kaydet(self):
        referans_turu, referans_id = self._secili_referans()
        if referans_turu != "serbest" and not referans_id:
            messagebox.showwarning("Uyarı", "Önce ilgili sporcu veya kulüp seçin.")
            return
        kalemler = []
        for kod, satir in self._evrak_satirlari.items():
            kalemler.append({
                "belge_kodu": kod,
                "belge_adi": satir["ad"],
                "zorunlu": satir["zorunlu"],
                "teslim": satir["teslim"].get(),
                "notlar": satir["notlar"].get().strip(),
            })
        db.evrak_kontrol_kaydet(
            self.v_basvuru.get(),
            referans_turu,
            kalemler,
            referans_id=referans_id,
            sezon=self.v_sezon.get().strip(),
        )
        self._ozet_listele()
        messagebox.showinfo("Başarılı", "Evrak kontrol kaydı kaydedildi.")

    def _referans_etiketi(self, referans_turu: str, referans_id: int) -> str:
        if referans_turu == "sporcu":
            for ad, sid in self._sporcu_map.items():
                if sid == referans_id:
                    return ad
            return f"Sporcu #{referans_id}"
        if referans_turu == "kulup":
            for ad, kid in self._kulup_map.items():
                if kid == referans_id:
                    return ad
            return f"Kulüp #{referans_id}"
        return "Genel"

    def _ozet_listele(self):
        self.tree_ozet.delete(*self.tree_ozet.get_children())
        self._ozet_map.clear()
        for index, row in enumerate(db.evrak_kontrol_ozet_listele()):
            iid = "|".join([
                row["basvuru_turu"],
                row["referans_turu"],
                str(row["referans_id"]),
                row["sezon"] or "",
            ])
            self._ozet_map[iid] = {
                "basvuru_turu": row["basvuru_turu"],
                "referans_turu": row["referans_turu"],
                "referans_id": row["referans_id"],
                "sezon": row["sezon"],
            }
            tag = "even" if index % 2 == 0 else "odd"
            self.tree_ozet.insert(
                "",
                "end",
                iid=iid,
                values=(
                    row["basvuru_turu"],
                    self._referans_etiketi(row["referans_turu"], row["referans_id"]),
                    row["sezon"] or "—",
                    f"{row['tamamlanan']}/{row['toplam']}",
                    row["guncelleme_tarihi"],
                ),
                tags=(tag,),
            )

    def _on_ozet_sec(self, _=None):
        sel = self.tree_ozet.selection()
        if not sel:
            return
        kayit = self._ozet_map.get(sel[0])
        if not kayit:
            return
        self.v_basvuru.set(kayit["basvuru_turu"])
        self.v_sezon.set(kayit["sezon"] or "")
        self._basvuru_degisti()
        if kayit["referans_turu"] == "sporcu":
            for ad, sid in self._sporcu_map.items():
                if sid == kayit["referans_id"]:
                    self.v_sporcu.set(ad)
                    break
        elif kayit["referans_turu"] == "kulup":
            for ad, kid in self._kulup_map.items():
                if kid == kayit["referans_id"]:
                    self.v_kulup.set(ad)
                    break
        self._mevcut_kaydi_yukle()

    def _temizle(self):
        self.v_sezon.set("2026")
        self._satirlari_sifirla()

    def _secili_referans_adi(self) -> str:
        referans_turu = self._aktif_kural()["referans_turu"]
        if referans_turu == "sporcu":
            return self.v_sporcu.get().strip() or "—"
        if referans_turu == "kulup":
            return self.v_kulup.get().strip() or "—"
        return "Genel"

    def _form_icerik_olustur(self) -> str:
        secim = self.v_form_ornek.get().strip()
        if not secim:
            raise ValueError("Seçili başvuru türü için talimat eki formu bulunmuyor.")
        ornek = self._aktif_form_map.get(secim)
        if not ornek:
            raise ValueError("Seçilen form örneği bulunamadı.")

        referans_adi = self._secili_referans_adi()
        sezon = self.v_sezon.get().strip() or "—"
        basvuru = self.v_basvuru.get().strip() or "—"
        bugun = date.today().strftime("%d.%m.%Y")

        if ornek["kod"] == "EK-7":
            satirlar = [
                "EK-7 VELİ MUVAFAKATNAME FORMU",
                "",
                "Tarih: ............................",
                "",
                "KKTC Bisiklet Federasyonu Başkanlığı'na,",
                "",
                "Velisi bulunduğum aşağıda bilgileri yer alan sporcunun, KKTC Bisiklet Federasyonu",
                "tarafından düzenlenen antrenman, yarış, kamp ve diğer sportif faaliyetlere katılmasına",
                "izin verdiğimi beyan ederim.",
                "",
                "Sporcunun Adı Soyadı : ..........................................................",
                "Kimlik No            : ..........................................................",
                "Doğum Tarihi         : ..........................................................",
                "Kulübü               : ..........................................................",
                "",
                "Velinin;",
                "",
                "Adı Soyadı : ....................................................................",
                "Telefon    : ....................................................................",
                "Adres      : ....................................................................",
                "",
                "İşbu muvafakatnameyi kendi rızam ile imzaladığımı kabul ederim.",
                "",
                "Veli / Yasal Temsilci",
                "İmza:",
            ]
            return "\n".join(satirlar)

        satirlar = [
            "KIBRIS TÜRK BİSİKLET FEDERASYONU",
            ornek["baslik"],
            "=" * 60,
            f"Tarih           : {bugun}",
            f"Sezon           : {sezon}",
            f"Başvuru Türü    : {basvuru}",
            f"Talimat Eki     : {ornek['kod']} ({ornek['ad']})",
            f"İlgili Kayıt    : {referans_adi}",
            "",
            "Teslim Evrakları:",
        ]

        for index, (_, belge_adi, zorunlu) in enumerate(self._aktif_kural()["belgeler"], start=1):
            tip = "Zorunlu" if zorunlu else "Opsiyonel"
            satirlar.append(f"{index:02d}. [ ] {belge_adi} ({tip})")

        satirlar.extend([
            "",
            "Açıklama:",
            "................................................................",
            "................................................................",
            "",
            "Başvuru Sahibi İmza : ____________________",
            "Kontrol Eden İmza   : ____________________",
        ])
        return "\n".join(satirlar)

    def _form_onizleme_ac(self):
        try:
            icerik = self._form_icerik_olustur()
        except ValueError as exc:
            messagebox.showwarning("Uyarı", str(exc))
            return

        wnd = tk.Toplevel(self)
        wnd.title("Form Önizleme")
        wnd.configure(bg=BG)
        wnd.geometry("860x680")

        txt = tk.Text(wnd, wrap="word", font=("Consolas", 11))
        txt.pack(fill="both", expand=True, padx=8, pady=(8, 4))
        txt.insert("1.0", icerik)

        alt = ttk.Frame(wnd)
        alt.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(alt, text="💾 TXT Olarak Kaydet", style="Neu.TButton",
                   command=lambda: self._formu_dosyaya_yaz(txt.get("1.0", "end-1c"), sor=True)).pack(side="left", padx=4)
        ttk.Button(alt, text="🖨 Yazdır", style="Neu.TButton",
                   command=lambda: self._formu_yazdir(txt.get("1.0", "end-1c"))).pack(side="left", padx=4)
        ttk.Button(alt, text="Kapat", style="Neu.TButton",
                   command=wnd.destroy).pack(side="right", padx=4)

    def _formu_dosyaya_yaz(self, icerik: str, sor: bool = False) -> str:
        if sor:
            from tkinter import filedialog
            yol = filedialog.asksaveasfilename(
                title="Formu kaydet",
                defaultextension=".txt",
                filetypes=[("Metin Dosyası", "*.txt"), ("Tüm Dosyalar", "*.*")],
                initialfile="evrak_form_ornegi.txt",
            )
            if not yol:
                return ""
        else:
            fd, yol = tempfile.mkstemp(prefix="ktbf_form_", suffix=".txt")
            os.close(fd)
        with open(yol, "w", encoding="utf-8") as dosya:
            dosya.write(icerik)
        self._son_form_dosyasi = yol
        return yol

    def _formu_yazdir(self, icerik: str | None = None):
        try:
            metin = icerik if icerik is not None else self._form_icerik_olustur()
        except ValueError as exc:
            messagebox.showwarning("Uyarı", str(exc))
            return

        try:
            yol = self._formu_dosyaya_yaz(metin, sor=False)
            if not yol:
                return
            if os.name == "nt":
                os.startfile(yol, "print")
                messagebox.showinfo("Yazdırma", "Form yazdırma komutu gönderildi.")
            else:
                messagebox.showinfo("Bilgi", f"Yazdırma yalnızca Windows'ta otomatik desteklenir.\nDosya: {yol}")
        except Exception as exc:
            messagebox.showerror("Hata", f"Form yazdırılamadı: {exc}")


class EvrakSablonPenceresi(tk.Toplevel):
    def __init__(self, parent: tk.Misc):
        super().__init__(parent)
        self.title("Evrak Şablonları")
        self.configure(bg=BG)
        self.geometry("1160x740")
        self.minsize(980, 620)

        self.v_basvuru_filtre = tk.StringVar(value="Tümü")
        self.v_sezon = tk.StringVar(value="2026")
        self.v_referans = tk.StringVar()
        self._sablon_map: dict[str, dict] = {}

        self._build()
        self._listeyi_yenile()

    def _build(self):
        ust = ttk.LabelFrame(self, text="Evrak Şablonları", padding=8)
        ust.pack(fill="x", padx=8, pady=(8, 4))

        ttk.Label(ust, text="Başvuru Filtresi").grid(
            row=0, column=0, sticky="e", padx=(8, 4), pady=4)
        cb = ttk.Combobox(
            ust,
            textvariable=self.v_basvuru_filtre,
            state="readonly",
            width=38,
            values=["Tümü", *EvrakKontrolSekme.BASVURU_TURLERI.keys()],
        )
        cb.grid(row=0, column=1, sticky="ew", padx=(0, 8), pady=4)
        cb.bind("<<ComboboxSelected>>", lambda _=None: self._listeyi_yenile())

        ttk.Label(ust, text="Sezon").grid(
            row=0, column=2, sticky="e", padx=(8, 4), pady=4)
        ttk.Entry(ust, textvariable=self.v_sezon, width=12).grid(
            row=0, column=3, sticky="w", padx=(0, 8), pady=4)

        ttk.Label(ust, text="İlgili Kayıt").grid(
            row=1, column=0, sticky="e", padx=(8, 4), pady=4)
        ttk.Entry(ust, textvariable=self.v_referans, width=44).grid(
            row=1, column=1, sticky="ew", padx=(0, 8), pady=4)

        btn = ttk.Frame(ust)
        btn.grid(row=1, column=2, columnspan=2, sticky="w", pady=4)
        ttk.Button(btn, text="🧾 Seçili Şablonu Üret", style="Neu.TButton",
                   command=self._seciliyi_onizle).pack(side="left", padx=4)
        ttk.Button(btn, text="💾 TXT Kaydet", style="Neu.TButton",
                   command=self._seciliyi_kaydet).pack(side="left", padx=4)
        ttk.Button(btn, text="🖨 Yazdır", style="Neu.TButton",
                   command=self._seciliyi_yazdir).pack(side="left", padx=4)

        ust.columnconfigure(1, weight=1)

        govde = ttk.PanedWindow(self, orient="horizontal")
        govde.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        sol = ttk.LabelFrame(govde, text="Talimat Eki Şablon Listesi", padding=4)
        govde.add(sol, weight=2)

        cols = [
            ("basvuru", "Başvuru Türü", 220),
            ("ek", "Ek Kodu", 70),
            ("form", "Form Adı", 240),
            ("kaynak", "Talimat Kaynağı", 260),
        ]
        tf, self.tree = _make_tree(sol, cols)
        tf.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_sec)

        sag = ttk.LabelFrame(govde, text="Şablon Önizleme", padding=4)
        govde.add(sag, weight=3)
        self.txt = tk.Text(sag, wrap="word", font=("Consolas", 11))
        self.txt.pack(fill="both", expand=True)

    def _listeyi_yenile(self):
        self.tree.delete(*self.tree.get_children())
        self._sablon_map.clear()
        filtre = self.v_basvuru_filtre.get().strip() or "Tümü"

        index = 0
        for basvuru, kural in EvrakKontrolSekme.BASVURU_TURLERI.items():
            if filtre != "Tümü" and basvuru != filtre:
                continue
            for form_kodu, form_adi, form_baslik in kural.get("formlar", []):
                iid = f"{basvuru}|{form_kodu}|{index}"
                self._sablon_map[iid] = {
                    "basvuru": basvuru,
                    "kaynak": kural.get("kaynak", "—"),
                    "form_kodu": form_kodu,
                    "form_adi": form_adi,
                    "form_baslik": form_baslik,
                    "belgeler": list(kural.get("belgeler", [])),
                }
                tag = "even" if index % 2 == 0 else "odd"
                self.tree.insert(
                    "",
                    "end",
                    iid=iid,
                    values=(basvuru, form_kodu, form_adi, kural.get("kaynak", "—")),
                    tags=(tag,),
                )
                index += 1

        secimler = self.tree.get_children()
        if secimler:
            self.tree.selection_set(secimler[0])
            self._seciliyi_onizle()
        else:
            self.txt.delete("1.0", "end")
            self.txt.insert("1.0", "Bu filtre için tanımlı şablon bulunamadı.")

    def _on_sec(self, _=None):
        self._seciliyi_onizle()

    def _secili_sablon(self):
        sec = self.tree.selection()
        if not sec:
            return None
        return self._sablon_map.get(sec[0])

    def _sablon_metni_uret(self, sablon: dict) -> str:
        bugun = date.today().strftime("%d.%m.%Y")
        referans = self.v_referans.get().strip() or "—"
        sezon = self.v_sezon.get().strip() or "—"

        if sablon["form_kodu"] == "EK-7":
            satirlar = [
                "EK-7 VELİ MUVAFAKATNAME FORMU",
                "",
                "Tarih: ............................",
                "",
                "KKTC Bisiklet Federasyonu Başkanlığı'na,",
                "",
                "Velisi bulunduğum aşağıda bilgileri yer alan sporcunun, KKTC Bisiklet Federasyonu",
                "tarafından düzenlenen antrenman, yarış, kamp ve diğer sportif faaliyetlere katılmasına",
                "izin verdiğimi beyan ederim.",
                "",
                "Sporcunun Adı Soyadı : ..........................................................",
                "Kimlik No            : ..........................................................",
                "Doğum Tarihi         : ..........................................................",
                "Kulübü               : ..........................................................",
                "",
                "Velinin;",
                "",
                "Adı Soyadı : ....................................................................",
                "Telefon    : ....................................................................",
                "Adres      : ....................................................................",
                "",
                "İşbu muvafakatnameyi kendi rızam ile imzaladığımı kabul ederim.",
                "",
                "Veli / Yasal Temsilci",
                "İmza:",
            ]
            return "\n".join(satirlar)

        satirlar = [
            "KIBRIS TÜRK BİSİKLET FEDERASYONU",
            sablon["form_baslik"],
            "=" * 68,
            f"Tarih           : {bugun}",
            f"Sezon           : {sezon}",
            f"Başvuru Türü    : {sablon['basvuru']}",
            f"Talimat Eki     : {sablon['form_kodu']} - {sablon['form_adi']}",
            f"Talimat Kaynağı : {sablon['kaynak']}",
            f"İlgili Kayıt    : {referans}",
            "",
            "Teslim Evrakları:",
        ]
        for i, (_, belge_adi, zorunlu) in enumerate(sablon["belgeler"], start=1):
            tip = "Zorunlu" if zorunlu else "Opsiyonel"
            satirlar.append(f"{i:02d}. [ ] {belge_adi} ({tip})")

        satirlar.extend([
            "",
            "Açıklama:",
            "................................................................",
            "................................................................",
            "",
            "Başvuru Sahibi İmza : ____________________",
            "Kontrol Eden İmza   : ____________________",
        ])
        return "\n".join(satirlar)

    def _seciliyi_onizle(self):
        sablon = self._secili_sablon()
        if not sablon:
            messagebox.showwarning("Uyarı", "Önce listeden bir şablon seçin.", parent=self)
            return
        metin = self._sablon_metni_uret(sablon)
        self.txt.delete("1.0", "end")
        self.txt.insert("1.0", metin)

    def _dosyaya_kaydet(self, icerik: str, sor: bool = False) -> str:
        if sor:
            from tkinter import filedialog
            yol = filedialog.asksaveasfilename(
                title="Evrak şablonunu kaydet",
                defaultextension=".txt",
                filetypes=[("Metin Dosyası", "*.txt"), ("Tüm Dosyalar", "*.*")],
                initialfile="evrak_sablonu.txt",
                parent=self,
            )
            if not yol:
                return ""
        else:
            fd, yol = tempfile.mkstemp(prefix="ktbf_evrak_sablon_", suffix=".txt")
            os.close(fd)
        with open(yol, "w", encoding="utf-8") as dosya:
            dosya.write(icerik)
        return yol

    def _seciliyi_kaydet(self):
        sablon = self._secili_sablon()
        if not sablon:
            messagebox.showwarning("Uyarı", "Önce listeden bir şablon seçin.", parent=self)
            return
        metin = self._sablon_metni_uret(sablon)
        yol = self._dosyaya_kaydet(metin, sor=True)
        if yol:
            messagebox.showinfo("Başarılı", f"Şablon dosyası kaydedildi.\n{yol}", parent=self)

    def _seciliyi_yazdir(self):
        sablon = self._secili_sablon()
        if not sablon:
            messagebox.showwarning("Uyarı", "Önce listeden bir şablon seçin.", parent=self)
            return
        try:
            metin = self._sablon_metni_uret(sablon)
            yol = self._dosyaya_kaydet(metin, sor=False)
            if not yol:
                return
            if os.name == "nt":
                os.startfile(yol, "print")
                messagebox.showinfo("Yazdırma", "Evrak şablonu yazdırma komutu gönderildi.", parent=self)
            else:
                messagebox.showinfo("Bilgi", f"Yazdırma yalnızca Windows'ta otomatik desteklenir.\nDosya: {yol}", parent=self)
        except Exception as exc:
            messagebox.showerror("Hata", f"Şablon yazdırılamadı: {exc}", parent=self)


# ---------------------------------------------------------------------------
# Geçmiş Sezon Görüntüleme Penceresi
# ---------------------------------------------------------------------------

class GecmisSezonPenceresi(tk.Toplevel):
    """temp/ klasöründeki Excel dosyalarından geçmiş sezon verilerini gösterir."""

    def __init__(self, parent: tk.Misc, okuyucu_modulu):
        super().__init__(parent)
        self._okuyucu = okuyucu_modulu
        self.title("📁 Geçmiş Sezon Verileri")
        self.configure(bg=BG)
        _apply_window_geometry(self, width_ratio=0.72, height_ratio=0.75,
                               min_width=800, min_height=560)

        sezonlar = self._okuyucu.mevcut_sezonlar()
        if not sezonlar:
            ttk.Label(self, text="Hiçbir Excel dosyası bulunamadı.\ntemp/ klasörünü kontrol edin.",
                      font=FONT).pack(expand=True)
            return

        # ── Üst kontroller ──────────────────────────────────────────────
        ust = ttk.Frame(self)
        ust.pack(fill="x", padx=8, pady=(8, 4))

        ttk.Label(ust, text="Sezon:").pack(side="left", padx=(0, 4))
        self.v_sezon = tk.StringVar(value=str(sezonlar[-1]))
        cb_sezon = ttk.Combobox(ust, textvariable=self.v_sezon,
                                values=[str(s) for s in sezonlar],
                                width=8, state="readonly")
        cb_sezon.pack(side="left", padx=4)
        cb_sezon.bind("<<ComboboxSelected>>", lambda _: self._listeyi_yenile())

        ttk.Separator(ust, orient="vertical").pack(side="left", fill="y", padx=8)

        self.v_goruntu = tk.StringVar(value="Sporcular")
        for deger, etiket in [("Sporcular", "🚴 Sporcular"),
                             ("Kulupler", "🏢 Kulüpler")]:
            rb = ttk.Radiobutton(ust, text=etiket, variable=self.v_goruntu,
                                 value=deger, command=self._listeyi_yenile)
            rb.pack(side="left", padx=4)

        ttk.Separator(ust, orient="vertical").pack(side="left", fill="y", padx=8)

        ttk.Label(ust, text="Sporcu adı:").pack(side="left", padx=(0, 4))
        self.v_sporcu_arama = tk.StringVar()
        self.v_sporcu_arama.trace_add("write", lambda *_: self._listeyi_yenile())
        ttk.Entry(ust, textvariable=self.v_sporcu_arama, width=22).pack(side="left", padx=4)

        self.lbl_sayac = ttk.Label(ust, text="", font=FONT_B)
        self.lbl_sayac.pack(side="left", padx=4)

        ttk.Button(ust, text="🔄 Yenile", style="Neu.TButton",
                   command=self._listeyi_yenile).pack(side="right", padx=4)
        ttk.Button(ust, text="📊 Excel'e Aktar", style="Neu.TButton",
                   command=self._excele_aktar).pack(side="right", padx=4)

        # ── Liste ───────────────────────────────────────────────────────
        cols_sporcu = [
            ("ad_soyad", "Ad Soyad", 200),
            ("lisans_no", "Lisans No", 110),
            ("kulup", "Kulüp", 200),
            ("cinsiyet", "Cinsiyet", 80),
            ("dogum", "Doğum", 160),
            ("telefon", "Telefon", 130),
        ]
        cols_kulup = [
            ("kulup_adi", "Kulüp Adı", 350),
            ("sporcu_sayisi", "Sporcu Sayısı", 140),
        ]
        self._cols_sporcu = cols_sporcu
        self._cols_kulup = cols_kulup

        tf, self.tree = _make_tree(self, cols_sporcu)
        tf.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        self._listeyi_yenile()

    def _listeyi_yenile(self):
        sezon = int(self.v_sezon.get())
        goruntu = self.v_goruntu.get()

        if goruntu == "Sporcular":
            self.tree.configure(columns=[c[0] for c in self._cols_sporcu],
                                show="headings")
            for cid, ctitle, cw in self._cols_sporcu:
                self.tree.heading(cid, text=ctitle)
                self.tree.column(cid, width=_scaled(cw), minwidth=_scaled(40))
            rows = self._okuyucu.sporcular_listele(sezon)
            arama = self.v_sporcu_arama.get().strip().casefold()
            if arama:
                rows = [r for r in rows if arama in r["ad_soyad"].casefold()]
            self.tree.delete(*self.tree.get_children())
            for i, r in enumerate(rows):
                tag = "even" if i % 2 == 0 else "odd"
                self.tree.insert("", "end", iid=str(i),
                                 values=(r["ad_soyad"], r["lisans_no"],
                                         r["kulup"], r["cinsiyet"],
                                         r["dogum"], r["telefon"]),
                                 tags=(tag,))
            self.lbl_sayac.config(text=f"{len(rows)} sporcu")
        else:
            self.tree.configure(columns=[c[0] for c in self._cols_kulup],
                                show="headings")
            for cid, ctitle, cw in self._cols_kulup:
                self.tree.heading(cid, text=ctitle)
                self.tree.column(cid, width=_scaled(cw), minwidth=_scaled(40))
            rows = self._okuyucu.kulupler_listele(sezon)
            self.tree.delete(*self.tree.get_children())
            for i, r in enumerate(rows):
                tag = "even" if i % 2 == 0 else "odd"
                self.tree.insert("", "end", iid=str(i),
                                 values=(r["kulup_adi"], r["sporcu_sayisi"]),
                                 tags=(tag,))
            self.lbl_sayac.config(text=f"{len(rows)} kulüp")

    def _excele_aktar(self):
        """Görüntülenen listeyi Excel olarak dışa aktarır."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from tkinter import filedialog

        sezon = int(self.v_sezon.get())
        goruntu = self.v_goruntu.get()
        rows = []
        basliklar = []

        if goruntu == "Sporcular":
            basliklar = ["Ad Soyad", "Lisans No", "Kulüp", "Cinsiyet", "Doğum", "Telefon"]
            for child in self.tree.get_children():
                vals = self.tree.item(child, "values")
                if vals:
                    rows.append(list(vals))
        else:
            basliklar = ["Kulüp Adı", "Sporcu Sayısı"]
            for child in self.tree.get_children():
                vals = self.tree.item(child, "values")
                if vals:
                    rows.append(list(vals))

        if not rows:
            messagebox.showinfo("Bilgi", "Dışa aktarılacak veri bulunamadı.", parent=self)
            return

        yol = filedialog.asksaveasfilename(
            title="Excel olarak kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx")],
            initialfile=f"{sezon}_{goruntu.lower()}.xlsx",
            parent=self,
        )
        if not yol:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"{sezon} {goruntu}"

            baslik_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
            baslik_font = Font(bold=True, size=11, color="FFFFFF")

            for col, b in enumerate(basliklar, 1):
                h = ws.cell(row=1, column=col, value=b)
                h.font = baslik_font
                h.fill = baslik_fill
                h.alignment = Alignment(horizontal="center")

            for i, row in enumerate(rows, start=2):
                for j, val in enumerate(row, 1):
                    ws.cell(row=i, column=j, value=val)

            wb.save(yol)
            messagebox.showinfo("Başarılı", f"Dışa aktarıldı:\n{yol}", parent=self)
        except Exception as exc:
            messagebox.showerror("Hata", f"Dışa aktarılamadı: {exc}", parent=self)


# ---------------------------------------------------------------------------
# Ana uygulama
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        # Sekmeler ilk açılışta sorgu yaptığı için DB önce hazır olmalı.
        db.init_db()
        self.title("KTBF – Lisans Kayıt Sistemi")
        _configure_display(self)
        self.configure(bg=BG)
        self.resizable(True, True)
        _style_widget(self)

        # Başlık
        hdr = tk.Frame(self, bg=HEADER, height=38)
        hdr.pack(fill="x")
        tk.Label(hdr,
                 text="  🚴 Kıbrıs Türk Bisiklet Federasyonu – Lisans Kayıt Sistemi",
                 bg=HEADER, fg="white",
             font=FONT_H).pack(side="left", pady=_scaled(4))
        ttk.Button(hdr, text="📁 Geçmiş Sezonlar", style="Neu.TButton",
                   command=self._gecmis_sezon_ac).pack(side="right", padx=4, pady=_scaled(4))
        ttk.Button(hdr, text="🌐 Anasayfa", style="Neu.TButton",
                   command=self._anasayfa_penceresi_ac).pack(side="right", padx=4, pady=_scaled(4))

        # Sekmeler
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=6, pady=6)

        self.kulup_sekme  = KulupSekme(self.nb)
        self.sporcu_sekme = SporcuSekme(self.nb)
        self.hib_sekme = HibSekme(self.nb)
        self.yaris_kayit_sekme = YarisKayitSekme(self.nb)

        self.nb.add(self.kulup_sekme,  text="  🏢 Kulüpler  ")
        self.nb.add(self.sporcu_sekme, text="  🚴 Sporcular  ")
        self.nb.add(self.hib_sekme, text="  🚲 HİB  ")
        self.nb.add(self.yaris_kayit_sekme, text="  🏁 Yarış Kayıt  ")

    def _sekme_ac(self, index: int):
        self.nb.select(index)

    def _local_server_calisiyor_mu(self) -> bool:
        try:
            with urllib.request.urlopen("http://127.0.0.1:8000/index.html", timeout=1.2) as yanit:
                return yanit.status == 200
        except Exception:
            return False

    def _local_server_baslat(self) -> bool:
        if self._local_server_calisiyor_mu():
            return True

        server_py = os.path.join(os.path.dirname(__file__), "local_server.py")
        if not os.path.exists(server_py):
            return False

        try:
            # Ayrı süreçte sessiz başlatılır; GUI kapanınca da çalışmaya devam eder.
            startupinfo = None
            creationflags = 0
            if os.name == "nt":
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)

            subprocess.Popen(
                [sys.executable, server_py, "--host", "127.0.0.1", "--port", "8000"],
                cwd=os.path.dirname(__file__),
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                startupinfo=startupinfo,
                creationflags=creationflags,
            )
        except Exception:
            return False

        # Server'ın ayağa kalkması için kısa bekleme/retry.
        for _ in range(12):
            if self._local_server_calisiyor_mu():
                return True
            time.sleep(0.12)
            self.update_idletasks()
        return False

    def _gecmis_sezon_ac(self):
        """Geçmiş sezon Excel dosyalarını görüntüleme penceresi açar."""
        try:
            import temp.excel_okuyucu as excel_okuyucu
        except ImportError:
            messagebox.showerror("Hata", "excel_okuyucu modülü bulunamadı.")
            return
        GecmisSezonPenceresi(self, excel_okuyucu)

    def _anasayfa_penceresi_ac(self):
        """Anasayfayı localhost üzerinden tarayıcıda aç."""
        try:
            if not self._local_server_baslat():
                messagebox.showerror(
                    "Hata",
                    "Local server başlatılamadı. local_server.py dosyasını ve port 8000 durumunu kontrol edin.",
                )
                return
            webbrowser.open("http://127.0.0.1:8000/index.html")
        except Exception as exc:
            messagebox.showerror("Hata", f"Tarayıcıda açılamadı: {exc}")




def main():
    _ensure_tcl_tk_env()
    app = App()
    try:
        app.mainloop()
    except KeyboardInterrupt:
        # Terminalden Ctrl+C ile kapatıldığında gereksiz traceback göstermeyelim.
        pass


if __name__ == "__main__":
    main()
